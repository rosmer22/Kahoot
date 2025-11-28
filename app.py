from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, g, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from io import BytesIO
from datetime import datetime, timedelta
from werkzeug.exceptions import HTTPException
from flask_mail import Mail, Message
import random
import math
import bd
import hashlib
import hmac
import base64
from flask_jwt import JWT, jwt_required, current_identity
try:
    import jwt as pyjwt
except ImportError:
    pyjwt = None
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import pickle
import logging
import os
import re
from flask_cors import CORS
from pathlib import Path  # ✅ agregado para rutas absolutas seguras
import json


def encriptar_sha256(cadena):
    cadbytes = cadena.encode('utf-8')
    sha256_hash_object = hashlib.sha256()
    sha256_hash_object.update(cadbytes)
    hex_digest = sha256_hash_object.hexdigest()
    return hex_digest

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app, resources={r"/*": {"origins": "*"}})
app.debug = True
# SECRET_KEY es necesario para Flask-JWT (igual que en Ejemplo.py)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')
app.config['SECRET_KEY'] = app.secret_key

# Configuración JWT (basada en Ejemplo.py pero con configs adicionales necesarias)
app.config['JWT_AUTH_USERNAME_KEY'] = 'email'
app.config['JWT_AUTH_PASSWORD_KEY'] = 'password'
app.config['JWT_AUTH_HEADER_PREFIX'] = 'JWT'
app.config['JWT_EXPIRATION_DELTA'] = timedelta(hours=12)
app.config['JWT_EXPIRATION_MINUTES'] = 12 * 60
app.config['JWT_COOKIE_NAME'] = 'access_token'
app.config['JWT_COOKIE_SECURE'] = False
app.config['JWT_COOKIE_SAMESITE'] = 'Lax'

# === Config generales (MODIFICADO: usar rutas absolutas y crear carpeta en carga) ===
BASE_DIR = Path(__file__).resolve().parent               # /home/usuario/mysite
STATIC_DIR = BASE_DIR / "static"
UPLOAD_DIR = STATIC_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)            # crea si no existe (funciona en WSGI)

app.config['UPLOAD_FOLDER'] = str(UPLOAD_DIR)            # ruta ABSOLUTA para guardar archivos

# === Config Google Drive OAuth ===
OAUTH_CONFIG_FILE = str(BASE_DIR / 'oauth_config.json')
TOKEN_FILE = str(BASE_DIR / 'token.pickle')
GOOGLE_DRIVE_FOLDER_ID = '1v1lgL9bQQMNPcfFFmvkHHo0KDpk5MOiV'
SCOPES = ['https://www.googleapis.com/auth/drive.file']

# Permitir HTTP en desarrollo local (solo para OAuth)
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

# === Helper: Obtener credenciales de Google Drive ===
def get_drive_credentials():
    """Obtiene o refresca las credenciales de Google Drive"""
    creds = None

    # Cargar token guardado si existe
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as token:
            creds = pickle.load(token)

    # Si no hay credenciales válidas, retorna None (se necesita autorizar)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                # Guardar token actualizado
                with open(TOKEN_FILE, 'wb') as token:
                    pickle.dump(creds, token)
            except Exception as e:
                # Si falla el refresh, necesita reautorizar
                return None
        else:
            return None

    return creds

# === Helper: Subir archivo a Google Drive ===
def subir_a_google_drive(file_stream, filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
    """
    Sube un archivo a Google Drive usando OAuth

    Args:
        file_stream: BytesIO con el contenido del archivo
        filename: Nombre del archivo
        mimetype: Tipo MIME del archivo

    Returns:
        dict: {'success': bool, 'file_id': str, 'file_url': str, 'needs_auth': bool} o error
    """
    try:
        # Obtener credenciales
        creds = get_drive_credentials()

        if not creds:
            return {
                'success': False,
                'needs_auth': True,
                'message': 'Se requiere autorización. Por favor, autoriza la aplicación primero.'
            }

        # Crear servicio de Drive
        service = build('drive', 'v3', credentials=creds)

        # Metadata del archivo
        file_metadata = {
            'name': filename,
            'parents': [GOOGLE_DRIVE_FOLDER_ID]
        }

        # Subir archivo
        media = MediaIoBaseUpload(file_stream, mimetype=mimetype, resumable=True)
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()

        return {
            'success': True,
            'needs_auth': False,
            'file_id': file.get('id'),
            'file_url': file.get('webViewLink'),
            'message': 'Archivo subido exitosamente a Google Drive'
        }

    except Exception as e:
        return {
            'success': False,
            'needs_auth': False,
            'error': str(e),
            'message': f'Error al subir a Google Drive: {str(e)}'
        }

from werkzeug.security import generate_password_hash
new_hash = generate_password_hash("Lostresconsejos@27", method='pbkdf2:sha256', salt_length=16)
print(new_hash)


# === Config de correo (Gmail SMTP) ===
# Sugerencia: usar variables de entorno para no exponer credenciales en código
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'valentinoandca@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', 'gcdl wgwf lego lmin')  # contraseña de aplicación
mail = Mail(app)

# Configuración básica de logging para registrar errores en la consola
logging.basicConfig(level=logging.ERROR)

# Almacén temporal de usuarios pendientes de verificación
# {email: {"username": str, "password": str, "codigo": int, "expira": datetime}}
usuarios_pendientes = {}
codigos = {}

from controllers import user_controller, quiz_controller


class User(object):
    def __init__(self, id, username, password, role='usuario'):
        self.id = id
        self.username = username
        self.password = password
        self.role = role

    def __str__(self):
        return "User(id='%s')" % self.id

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'role': self.role
        }


class UserIdentity(object):
    def __init__(self, user_dict):
        self.id = user_dict.get('id')
        self.username = user_dict.get('username')
        self.email = user_dict.get('email')
        self.role = user_dict.get('role', 'usuario')

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'email': self.email,
            'role': self.role
        }


def authenticate(username, password):
    try:
        import sys
        print(f"\n=== AUTHENTICATE LLAMADO ===", file=sys.stderr)
        print(f"Email: {username}", file=sys.stderr)
        sys.stderr.flush()

        usuario = user_controller.obtener_usuario_por_email(username)
        if not usuario:
            print(f"Usuario no encontrado para email: {username}", file=sys.stderr)
            sys.stderr.flush()
            return None

        print(f"Usuario encontrado en BD: {usuario.get('id')}", file=sys.stderr)
        sys.stderr.flush()

        # Comparar password hasheado
        if usuario['password'] == encriptar_sha256(password):
            user_obj = User(usuario["id"], usuario["email"], usuario["password"], usuario["role"])
            print(f"AUTENTICACION EXITOSA - ID: {user_obj.id}, Username: {user_obj.username}, Role: {user_obj.role}", file=sys.stderr)
            sys.stderr.flush()
            return user_obj

        print(f"Password no coincide para usuario: {username}", file=sys.stderr)
        sys.stderr.flush()
        return None
    except Exception as e:
        import sys
        print(f"ERROR en authenticate: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.stderr.flush()
        return None


def identity(payload):
    user_id = payload['identity']
    usuario = user_controller.obtener_usuario_por_id(user_id)
    user = None
    if usuario:
        user = User(usuario["id"], usuario["email"], usuario["password"], usuario["role"])
    return user



# Inicializar Flask-JWT (configuración simple como en Ejemplo.py)
jwt_manager = JWT(app, authenticate, identity)

# SOLO personalizar la respuesta para que devuelva JSON en vez de HTML
@jwt_manager.auth_response_handler
def custom_auth_response(access_token, identity):
    """Devuelve JSON en vez de HTML"""
    import sys
    try:
        print(f"\n=== AUTH RESPONSE HANDLER LLAMADO ===", file=sys.stderr)
        print(f"Token recibido: {str(access_token)[:30]}...", file=sys.stderr)
        print(f"Identity type: {type(identity)}", file=sys.stderr)
        print(f"Identity value: {identity}", file=sys.stderr)

        # Intentar acceder a los atributos
        try:
            user_id = identity.id
            username = identity.username
            role = identity.role
            print(f"Atributos obtenidos: id={user_id}, username={username}, role={role}", file=sys.stderr)
        except Exception as e:
            print(f"ERROR al acceder atributos: {e}", file=sys.stderr)
            # Si identity no tiene los atributos esperados, intentar acceder de otra forma
            user_id = getattr(identity, 'id', None)
            username = getattr(identity, 'username', None)
            role = getattr(identity, 'role', 'usuario')
            print(f"Atributos con getattr: id={user_id}, username={username}, role={role}", file=sys.stderr)

        sys.stderr.flush()

        response = jsonify({
            'access_token': access_token,
            'user': {
                'id': user_id,
                'username': username,
                'role': role
            }
        })

        print(f"Respuesta JSON creada exitosamente", file=sys.stderr)
        sys.stderr.flush()
        return response

    except Exception as e:
        print(f"\n!!! ERROR CRITICO en auth_response_handler !!!", file=sys.stderr)
        print(f"Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.stderr.flush()

        # Intentar devolver algo aunque sea
        return jsonify({
            'error': 'Error al generar respuesta',
            'message': str(e),
            'access_token': str(access_token) if access_token else None
        }), 500


@jwt_manager.jwt_error_handler
def custom_jwt_error_handler(error):
    """Maneja errores de JWT"""
    import sys
    print(f"\n=== JWT ERROR HANDLER ===", file=sys.stderr)
    print(f"Error: {str(error)}", file=sys.stderr)
    sys.stderr.flush()

    return jsonify({
        'error': 'Authentication failed',
        'message': str(error)
    }), 401


# Funciones helper para JWT (necesarias para endpoints personalizados)
def _b64url_encode(data: bytes) -> bytes:
    return base64.urlsafe_b64encode(data).rstrip(b'=')


def _b64url_decode(data: str) -> bytes:
    padding = '=' * (-len(data) % 4)
    return base64.urlsafe_b64decode(data + padding)


def build_jwt_payload(identity_obj):
    """Construye el payload del token JWT"""
    from datetime import datetime, timedelta
    now = datetime.utcnow()
    exp_delta = timedelta(minutes=app.config['JWT_EXPIRATION_MINUTES'])
    payload = {
        'identity': identity_obj.id,
        'username': identity_obj.username,
        'role': identity_obj.role,
        'iat': int(now.timestamp()),
        'nbf': int(now.timestamp()),
        'exp': int((now + exp_delta).timestamp())
    }
    return payload


def _encode_jwt_payload(payload: dict) -> str:
    """Codifica un payload a token JWT"""
    secret = app.config['SECRET_KEY']
    algorithm = 'HS256'

    # Intentar usar PyJWT si está disponible
    if pyjwt and hasattr(pyjwt, 'encode'):
        token = pyjwt.encode(payload, secret, algorithm=algorithm)
        if isinstance(token, bytes):
            token = token.decode('utf-8')
        return token

    # Fallback manual si PyJWT no disponible
    header = {'typ': 'JWT', 'alg': algorithm}
    header_bytes = json.dumps(header, separators=(',', ':'), sort_keys=True).encode('utf-8')
    payload_bytes = json.dumps(payload, separators=(',', ':'), sort_keys=True).encode('utf-8')
    segments = [_b64url_encode(header_bytes), _b64url_encode(payload_bytes)]
    signing_input = b'.'.join(segments)
    signature = hmac.new(secret.encode('utf-8'), signing_input, hashlib.sha256).digest()
    segments.append(_b64url_encode(signature))
    return '.'.join(segment.decode('utf-8') for segment in segments)


def _decode_jwt_payload(token: str) -> dict:
    """Decodifica un token JWT"""
    secret = app.config['SECRET_KEY']
    algorithm = 'HS256'

    # Intentar usar PyJWT si está disponible
    if pyjwt and hasattr(pyjwt, 'decode'):
        return pyjwt.decode(token, secret, algorithms=[algorithm])

    # Fallback manual
    try:
        header_b64, payload_b64, signature_b64 = token.split('.')
    except ValueError as exc:
        raise ValueError('Token JWT mal formado') from exc
    signing_input = f'{header_b64}.{payload_b64}'.encode('utf-8')
    signature = _b64url_decode(signature_b64)
    expected_signature = hmac.new(secret.encode('utf-8'), signing_input, hashlib.sha256).digest()
    if not hmac.compare_digest(signature, expected_signature):
        raise ValueError('Firma JWT inválida')
    payload = json.loads(_b64url_decode(payload_b64))
    exp = payload.get('exp')
    if exp is not None and int(exp) < int(datetime.utcnow().timestamp()):
        raise ValueError('Token JWT expirado')
    return payload


def generar_token_usuario(usuario):
    if not isinstance(usuario, UserIdentity):
        identidad = UserIdentity(usuario)
    else:
        identidad = usuario
    payload = build_jwt_payload(identidad)
    token = _encode_jwt_payload(payload)
    if isinstance(token, bytes):
        token = token.decode('utf-8')
    return token


def obtener_token_actual():
    auth_header = request.headers.get('Authorization', '')
    token = None
    prefix = app.config['JWT_AUTH_HEADER_PREFIX']
    if auth_header.startswith(f'{prefix} '):
        token = auth_header.split(' ', 1)[1].strip()
    elif auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    if not token:
        token = request.cookies.get(app.config['JWT_COOKIE_NAME'])
    return token


def adjuntar_token_response(response, token):
    if isinstance(token, UserIdentity):
        token = generar_token_usuario(token)
    if not isinstance(token, str):
        token = str(token)
    response.set_cookie(
        app.config['JWT_COOKIE_NAME'],
        token,
        max_age=app.config['JWT_EXPIRATION_MINUTES'] * 60,
        httponly=True,
        secure=app.config['JWT_COOKIE_SECURE'],
        samesite=app.config['JWT_COOKIE_SAMESITE']
    )
    return response


def limpiar_token_response(response):
    response.set_cookie(
        app.config['JWT_COOKIE_NAME'],
        '',
        max_age=0,
        expires=0,
        httponly=True,
        secure=app.config['JWT_COOKIE_SECURE'],
        samesite=app.config['JWT_COOKIE_SAMESITE']
    )
    return response


def _obtener_usuario_desde_token(token):
    if not token:
        return None
    try:
        payload = _decode_jwt_payload(token)
        identidad = identity(payload)
        if identidad:
            return identidad.to_dict()
    except Exception as exc:
        app.logger.info(f'Token JWT inválido o expirado: {exc}')
    return None



# =====================
#  Hooks / Helpers
# =====================
@app.before_request
def load_logged_in_user():
    token = obtener_token_actual()
    if token and 'HTTP_AUTHORIZATION' not in request.environ:
        request.environ['HTTP_AUTHORIZATION'] = f"{app.config['JWT_AUTH_HEADER_PREFIX']} {token}"
    g.user = _obtener_usuario_desde_token(token)

def is_auth():
    return g.user is not None

@app.context_processor
def inject_globals():
    return dict(is_auth=is_auth(), user=getattr(g, 'user', None))


# =====================
#  Sistema de Recompensas (Puntosmoneda)
# =====================
def asignar_puntosmoneda_grupo(sesion_id):
    """
    Calcula el ranking final de una sesión de grupo y asigna puntosmoneda:
    - 1er lugar: 300 puntosmoneda
    - 2do lugar: 150 puntosmoneda
    - 3er lugar: 75 puntosmoneda
    """
    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # ⚠️ IMPORTANTE: Verificar si ya se asignaron recompensas para esta sesión
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM historial_recompensas
            WHERE sesion_id = %s AND tipo_sesion = 'grupo'
        """, (sesion_id,))

        ya_asignadas = cursor.fetchone()['total']

        if ya_asignadas > 0:
            app.logger.info(f"Recompensas ya asignadas para sesión grupo {sesion_id}, saltando...")
            return  # Ya se asignaron, no hacer nada

        # Obtener el ranking de jugadores (excluyendo al creador)
        cursor.execute("""
            SELECT
                rg.user_id,
                u.username,
                SUM(rg.puntos) as puntos_totales,
                sg.iniciado_por
            FROM respuestas_grupo rg
            JOIN users u ON rg.user_id = u.id
            JOIN sesiones_grupo sg ON rg.sesion_id = sg.id
            WHERE rg.sesion_id = %s
            GROUP BY rg.user_id, u.username, sg.iniciado_por
            HAVING rg.user_id != sg.iniciado_por
            ORDER BY puntos_totales DESC
            LIMIT 3
        """, (sesion_id,))

        ranking = cursor.fetchall()

        if not ranking:
            return

        # Asignar recompensas según posición
        recompensas = {
            0: 300,  # 1er lugar
            1: 150,   # 2do lugar
            2: 75    # 3er lugar
        }

        for posicion, jugador in enumerate(ranking):
            if posicion < 3:  # Solo primeros 3 lugares
                puntosmoneda = recompensas.get(posicion, 0)
                user_id = jugador['user_id']
                puntos_totales = jugador['puntos_totales']

                # Actualizar puntosmoneda del usuario
                cursor.execute("""
                    UPDATE users
                    SET puntosmoneda = puntosmoneda + %s
                    WHERE id = %s
                """, (puntosmoneda, user_id))

                # Registrar en historial de recompensas
                cursor.execute("""
                    INSERT INTO historial_recompensas
                    (user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales)
                    VALUES (%s, %s, 'grupo', %s, %s, %s)
                """, (user_id, sesion_id, posicion + 1, puntosmoneda, puntos_totales))

        db.commit()
        app.logger.info(f"✅ Recompensas asignadas correctamente para sesión grupo {sesion_id}")

    except Exception as e:
        if db:
            db.rollback()
        app.logger.error(f"Error al asignar puntosmoneda en sesión {sesion_id}: {str(e)}")

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


def asignar_puntosmoneda_individual(sesion_id):
    """
    Calcula el ranking final de una sesión individual y asigna puntosmoneda:
    - 1er lugar: 300 puntosmoneda
    - 2do lugar: 150 puntosmoneda
    - 3er lugar: 75 puntosmoneda
    """
    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # ⚠️ IMPORTANTE: Verificar si ya se asignaron recompensas para esta sesión
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM historial_recompensas
            WHERE sesion_id = %s AND tipo_sesion = 'individual'
        """, (sesion_id,))

        ya_asignadas = cursor.fetchone()['total']

        if ya_asignadas > 0:
            app.logger.info(f"Recompensas ya asignadas para sesión individual {sesion_id}, saltando...")
            return  # Ya se asignaron, no hacer nada

        # Obtener el ranking de jugadores (excluyendo al creador)
        cursor.execute("""
            SELECT
                ri.user_id,
                u.username,
                SUM(ri.puntos) as puntos_totales,
                si.iniciado_por
            FROM respuestas_individual ri
            JOIN users u ON ri.user_id = u.id
            JOIN sesiones_individual si ON ri.sesion_id = si.id
            WHERE ri.sesion_id = %s
            GROUP BY ri.user_id, u.username, si.iniciado_por
            HAVING ri.user_id != si.iniciado_por
            ORDER BY puntos_totales DESC
            LIMIT 3
        """, (sesion_id,))

        ranking = cursor.fetchall()

        if not ranking:
            return

        # Asignar recompensas según posición
        recompensas = {
            0: 300,  # 1er lugar
            1: 150,   # 2do lugar
            2: 75    # 3er lugar
        }

        for posicion, jugador in enumerate(ranking):
            if posicion < 3:  # Solo primeros 3 lugares
                puntosmoneda = recompensas.get(posicion, 0)
                user_id = jugador['user_id']
                puntos_totales = jugador['puntos_totales']

                # Actualizar puntosmoneda del usuario
                cursor.execute("""
                    UPDATE users
                    SET puntosmoneda = puntosmoneda + %s
                    WHERE id = %s
                """, (puntosmoneda, user_id))

                # Registrar en historial de recompensas
                cursor.execute("""
                    INSERT INTO historial_recompensas
                    (user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales)
                    VALUES (%s, %s, 'individual', %s, %s, %s)
                """, (user_id, sesion_id, posicion + 1, puntosmoneda, puntos_totales))

        db.commit()
        app.logger.info(f"✅ Recompensas asignadas correctamente para sesión individual {sesion_id}")

    except Exception as e:
        if db:
            db.rollback()
        app.logger.error(f"Error al asignar puntosmoneda en sesión individual {sesion_id}: {str(e)}")

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


# =====================
#  Rutas OAuth Google Drive
# =====================
@app.route('/authorize-drive')
def authorize_drive():
    """Inicia el proceso de autorización OAuth para Google Drive"""
    if not is_auth():
        flash('Debes iniciar sesión primero', 'error')
        return redirect(url_for('login'))

    try:
        # Generar redirect_uri
        redirect_uri = url_for('oauth2callback', _external=True)
        print(f"\n🔍 DEBUG: Redirect URI = {redirect_uri}\n")

        # Crear flujo de OAuth
        flow = Flow.from_client_secrets_file(
            OAUTH_CONFIG_FILE,
            scopes=SCOPES,
            redirect_uri=redirect_uri
        )

        # Generar URL de autorización
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )

        # Guardar state en sesión para verificación
        session['oauth_state'] = state

        return redirect(authorization_url)

    except Exception as e:
        flash(f'Error al iniciar autorización: {str(e)}', 'error')
        return redirect(url_for('home'))

@app.route('/oauth2callback')
def oauth2callback():
    """Callback de OAuth después de la autorización"""
    if not is_auth():
        flash('Debes iniciar sesión primero', 'error')
        return redirect(url_for('login'))

    try:
        # Verificar state para prevenir CSRF
        state = session.get('oauth_state')

        # Crear flujo de OAuth
        flow = Flow.from_client_secrets_file(
            OAUTH_CONFIG_FILE,
            scopes=SCOPES,
            state=state,
            redirect_uri=url_for('oauth2callback', _external=True)
        )

        # Intercambiar código por credenciales
        flow.fetch_token(authorization_response=request.url)

        # Guardar credenciales
        creds = flow.credentials
        with open(TOKEN_FILE, 'wb') as token:
            pickle.dump(creds, token)

        flash('¡Autorización exitosa! Ahora puedes exportar resultados a Google Drive.', 'success')
        return redirect(url_for('home'))

    except Exception as e:
        flash(f'Error durante la autorización: {str(e)}', 'error')
        return redirect(url_for('home'))

@app.route('/check-drive-auth')
def check_drive_auth():
    """Verifica si la aplicación está autorizada para Google Drive"""
    if not is_auth():
        return jsonify({'authorized': False, 'message': 'Usuario no autenticado'})

    creds = get_drive_credentials()
    return jsonify({
        'authorized': creds is not None,
        'message': 'Autorización válida' if creds else 'Se requiere autorización'
    })


# =====================
#  Rutas públicas
# =====================
@app.route('/')
def home():
    """Página de inicio que muestra los cuestionarios públicos más recientes."""
    db = bd.obtener_conexion()
    cursor = db.cursor()
    cursor.execute(
        """
        SELECT
            c.id,
            COALESCE(c.titulo, 'Sin título') as titulo,
            c.pin,
            c.imagen_portada,
            (SELECT COUNT(*) FROM preguntas p WHERE p.cuestionario_id = c.id) as question_count
        FROM cuestionarios c
        WHERE c.estado = 'publico'
        ORDER BY c.created_at DESC
        LIMIT 6
        """
    )
    quizzes = cursor.fetchall()
    cursor.close()
    db.close()
    return render_template('home.html', title='RoBot', quizzes=quizzes)



@app.errorhandler(404)
def page_not_found(e):
    """Maneja errores 404 (Página no encontrada)."""
    return render_template('error.html', title='Página no encontrada'), 404

@app.errorhandler(Exception)
def handle_exception(e):
    """Maneja excepciones no capturadas para evitar que la aplicación se caiga."""
    # Si es una excepción HTTP estándar (como 404, 401, etc.), deja que Flask la maneje.
    if isinstance(e, HTTPException):
        return e

    # Para cualquier otra excepción (errores 500), regístrala
    import sys
    import traceback

    # Obtener mensaje de error de forma segura sin usar str()
    try:
        error_message = repr(e)
    except:
        error_message = "Error desconocido"

    print(f"\n!!! ERROR 500 CAPTURADO !!!", file=sys.stderr)
    print(f"Path: {request.path}", file=sys.stderr)
    print(f"Error: {error_message}", file=sys.stderr)
    traceback.print_exc(file=sys.stderr)
    sys.stderr.flush()

    # Si es una ruta API o /auth, devolver JSON en lugar de HTML
    if request.path.startswith('/api/') or request.path == '/auth':
        return jsonify({
            'success': False,
            'error': 'Error interno del servidor',
            'message': error_message if app.debug else 'Ha ocurrido un error inesperado'
        }), 500

    # Para rutas normales, mostrar la página de error
    flash("Ha ocurrido un error inesperado en el servidor. Nuestro equipo ha sido notificado.", "error")
    return render_template('error.html', title='Error del Sistema'), 500


@app.route('/empezar', methods=['GET'])
def empezar():
    if not g.user:
        flash('Debes iniciar sesión para empezar un cuestionario', 'warning')
        return redirect(url_for('login'))

    # Solo docentes pueden empezar cuestionarios
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden empezar cuestionarios', 'error')
        return redirect(url_for('home'))

    return render_template('empezar.html', title='Empezar')



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = user_controller.obtener_usuario_por_email(email)

        if not user:
            flash('No existe una cuenta con ese usuario o correo. Por favor, regístrate.', 'error')
            return redirect(url_for('login'))

        if user['password'] == encriptar_sha256(password):
            token = generar_token_usuario(user)
            flash('¡Bienvenido de nuevo!', 'success')
            response = redirect(url_for('home'))
            response.set_cookie('user_id', encriptar_sha256(str(user['id'])))
            response.set_cookie('username', encriptar_sha256(user['username']))
            return adjuntar_token_response(response, token)
        else:
            flash('Usuario o contraseña incorrectos.', 'error')
            return redirect(url_for('login'))

    return render_template('login.html', title='Iniciar Sesión')


@app.route('/logout')
def logout():
    session.clear()
    response = redirect(url_for('home'))
    response.set_cookie('user_id', '', max_age=0, expires=0)
    response.set_cookie('username', '', max_age=0, expires=0)
    return limpiar_token_response(response)


# =====================
#  Registro con verificación por correo
# =====================
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email').strip()
        username = request.form.get('username').strip()
        password = request.form.get('password')

        # Validaciones previas
        if user_controller.obtener_usuario_por_username(username):
            flash('El nombre de usuario ya existe', 'error')
            return redirect(url_for('register'))

        if user_controller.obtener_usuario_por_email(email):
            flash('El correo electrónico ya está en uso', 'error')
            return redirect(url_for('register'))

        if not (email.endswith('@usat.edu.pe') or email.endswith('@usat.pe')):
            flash('El correo debe pertenecer al dominio usat.edu.pe o usat.pe', 'error')
            return redirect(url_for('register'))
        # validación de minuscula, mayuscula, numero y caracter
        pattern = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&#._-])[A-Za-z\d@$!%*?&#._-]{8,}$'

        if not re.match(pattern, password):
            flash('La contraseña debe tener al menos 8 caracteres, incluir una mayúscula, una minúscula, un número y un carácter especial.', 'error')
            return redirect(url_for('register'))

        # Generar código de verificación
        codigo = random.randint(100000, 999999)
        expira = datetime.now() + timedelta(minutes=10)

        usuarios_pendientes[email] = {
            'username': username,
            'password': password,
            'codigo': codigo,
            'expira': expira,
        }

        # Enviar correo con el código (si hay credenciales válidas)
        try:
            msg = Message(
                'Código de confirmación - RoBot',
                sender=app.config['MAIL_USERNAME'],
                recipients=[email]
            )
            msg.body = f"Tu código de confirmación es: {codigo}\nVálido por 10 minutos."
            mail.send(msg)
            session['email_verificacion'] = email
            flash('Se ha enviado un código de verificación a tu correo. Revisa tu bandeja.', 'info')
            return redirect(url_for('verify_email'))
        except Exception as e:
            # Si el correo falla, se puede permitir registro directo o mostrar error.
            # Aquí optamos por mostrar error para no crear cuentas sin verificación.
            flash(f'No se pudo enviar el correo de verificación: {e}', 'error')
            return redirect(url_for('register'))

    return render_template('register.html', title='Registrarme')

import sys

@app.route('/codigo_recuperacion', methods=['GET', 'POST'])
def codigo_recuperacion():
    if request.method == 'POST':
        email = request.form.get('email').strip()

        if user_controller.obtener_usuario_por_email(email):
            # Generar código de 6 dígitos
            codigo = random.randint(100000, 999999)
            print(f"[DEBUG] Código de verificación para {email}: {codigo}")
            sys.stdout.flush()


            codigos[email] = {
                'codigo': str(codigo)
            }

            try:
                msg = Message(
                    'Código de confirmación - RoBot',
                    sender=app.config['MAIL_USERNAME'],
                    recipients=[email]
                )
                msg.body = f"Tu código de confirmación es: {codigo}\n(Sin expiración por ahora)."
                mail.send(msg)
                session['email_verificacion'] = email
                flash('Se ha enviado un código de verificación a tu correo. Revisa tu bandeja.', 'info')
                # Si tu formulario de código está en esta ruta, está bien:
                return redirect(url_for('verificar_codigo_recuperacion'))
            except Exception as e:
                print(f"[ERROR] No se pudo enviar el correo de verificación: {e}")
                sys.stdout.flush()
                flash('No se pudo enviar el correo de verificación. Usa el código que aparece en el log.', 'error')
                return redirect(url_for('verificar_codigo_recuperacion'))
        else:
            flash('El correo electrónico no existe', 'error')
            return redirect(url_for('recuperar_cuenta'))

    # Si entras por GET, por si acaso lo mandamos al form de recuperar
    return redirect(url_for('recuperar_cuenta'))


@app.route('/verify_email', methods=['GET', 'POST'])
def verify_email():
    email = session.get('email_verificacion')
    if not email:
        return redirect(url_for('register'))

    # Validar si el correo ya pertenece a un usuario registrado
    if user_controller.obtener_usuario_por_email(email):
        session.pop('email_verificacion', None)
        flash('Ya tienes una cuenta con este correo. Por favor, inicia sesión.', 'info')
        return redirect(url_for('login'))

    if request.method == 'POST':
        codigo_ingresado = request.form.get('codigo')
        datos = usuarios_pendientes.get(email)

        if datos and str(datos['codigo']) == codigo_ingresado and datetime.now() < datos['expira']:
            # Crear usuario definitivo
            user_controller.insertar_usuario(datos['username'], email, datos['password'])

            # Recuperar el usuario recién creado
            nuevo_usuario = user_controller.obtener_usuario_por_email(email)

            # Eliminar datos temporales
            usuarios_pendientes.pop(email, None)
            session.pop('email_verificacion', None)

            flash('Cuenta verificada y creada correctamente ✅', 'success')

            # Iniciar sesión automáticamente con la nueva cuenta
            if nuevo_usuario:
                token = generar_token_usuario(nuevo_usuario)
                flash('¡Bienvenido!', 'success')
                response = redirect(url_for('home'))
                response.set_cookie('user_id', encriptar_sha256(str(nuevo_usuario['id'])))
                response.set_cookie('username', encriptar_sha256(nuevo_usuario['username']))
                return adjuntar_token_response(response, token)
            else:
                flash('Error al iniciar sesión automáticamente. Intenta iniciar sesión manualmente.', 'warning')
                return redirect(url_for('login'))
        else:
            flash('Código inválido o expirado ❌', 'error')

    return render_template('emailverificacion.html', title='Verificar correo')

@app.route('/verificar_codigo_recuperacion', methods=['GET', 'POST'])
def verificar_codigo_recuperacion():
    email = session.get('email_verificacion')
    if not email:
        flash('Sesión inválida. Vuelve a solicitar un código.', 'error')
        return redirect(url_for('recuperar_cuenta'))

    if request.method == 'POST':
        codigo_ingresado = request.form.get('codigo', '').strip()
        datos = codigos.get(email)

        if datos and datos.get('codigo') == codigo_ingresado:
            codigos.pop(email, None)
            return redirect(url_for('cambiar_contra'))
        else:
            flash('Código inválido', 'error')

    # ⬇⬇⬇ CAMBIA ESTA LÍNEA
    return render_template('verificar_recuperar_cuenta.html', title='Verificar código')



@app.route('/cambiar_contra', methods=['GET'])
def cambiar_contra():
     return render_template('cambiar_contra.html', title='Empezar')

@app.route('/resend_code', methods=['POST'])
def resend_verification_code():
    """Reenvía un nuevo código de verificación al correo en sesión."""
    email = session.get('email_verificacion')
    if not email or email not in usuarios_pendientes:
        return jsonify({'success': False, 'message': 'No hay una verificación pendiente o la sesión ha expirado.'}), 400

    # Generar nuevo código y actualizar datos
    codigo = random.randint(100000, 999999)
    expira = datetime.now() + timedelta(minutes=10)

    usuarios_pendientes[email]['codigo'] = codigo
    usuarios_pendientes[email]['expira'] = expira

    # Enviar correo con el nuevo código
    try:
        msg = Message(
            'Nuevo Código de confirmación - RoBot',
            sender=app.config['MAIL_USERNAME'],
            recipients=[email]
        )
        msg.body = f"Tu nuevo código de confirmación es: {codigo}\nVálido por 10 minutos."
        mail.send(msg)
        return jsonify({'success': True, 'message': 'Se ha enviado un nuevo código a tu correo.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'No se pudo reenviar el correo: {e}'}), 500


# =====================
#  Perfil y seguridad
# =====================
@app.route('/settings')
def settings():
    if g.user is None:
        return redirect(url_for('login'))
    return render_template('settings.html', title='Configuración', user=g.user)

@app.route('/mis-puntos')
def mis_puntos():
    """Página para ver los puntosmoneda y rango del usuario"""
    if g.user is None:
        flash('Debes iniciar sesión para ver tus puntos', 'warning')
        return redirect(url_for('login'))

    # Verificar si el usuario es docente
    es_docente = g.user.get('role') == 'docente'

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Si es docente, mostrar ranking de todos los alumnos
        if es_docente:
            cursor.execute("""
                SELECT
                    u.id,
                    u.username,
                    u.email,
                    u.puntosmoneda,
                    r.nombre as rango_nombre,
                    r.icono as rango_icono,
                    r.color as rango_color
                FROM users u
                LEFT JOIN rangos r ON u.puntosmoneda >= r.puntos_minimos
                                   AND u.puntosmoneda <= r.puntos_maximos
                WHERE u.role = 'alumno'
                ORDER BY u.puntosmoneda DESC
            """)

            ranking_alumnos = cursor.fetchall()

            # Calcular puntos totales de todos los alumnos
            cursor.execute("""
                SELECT SUM(puntosmoneda) as total_puntos
                FROM users
                WHERE role = 'alumno'
            """)

            total_result = cursor.fetchone()
            total_puntos_alumnos = total_result['total_puntos'] if total_result else 0

            return render_template('mis_puntos.html',
                                 title='Ranking de Alumnos',
                                 es_docente=True,
                                 ranking_alumnos=ranking_alumnos,
                                 total_puntos_alumnos=total_puntos_alumnos)

        # Si es alumno, mostrar sus puntos individuales
        # Obtener información del usuario con su rango
        cursor.execute("""
            SELECT
                u.id,
                u.username,
                u.email,
                u.puntosmoneda,
                r.nombre as rango_nombre,
                r.icono as rango_icono,
                r.color as rango_color,
                r.orden as rango_orden,
                r.puntos_minimos,
                r.puntos_maximos,
                CASE
                    WHEN r.puntos_maximos >= 99999999 THEN 0
                    ELSE (r.puntos_maximos + 1 - u.puntosmoneda)
                END as puntos_para_siguiente_rango,
                (SELECT nombre
                 FROM rangos r2
                 WHERE r2.orden = r.orden + 1
                 LIMIT 1) as siguiente_rango,
                (SELECT icono
                 FROM rangos r2
                 WHERE r2.orden = r.orden + 1
                 LIMIT 1) as siguiente_rango_icono
            FROM users u
            LEFT JOIN rangos r ON u.puntosmoneda >= r.puntos_minimos
                               AND u.puntosmoneda <= r.puntos_maximos
            WHERE u.id = %s
        """, (g.user['id'],))

        usuario_info = cursor.fetchone()

        # Obtener todos los rangos para mostrar el progreso
        cursor.execute("""
            SELECT nombre, puntos_minimos, puntos_maximos, icono, color, orden
            FROM rangos
            ORDER BY orden
        """)

        todos_rangos = cursor.fetchall()

        # Obtener historial de recompensas del usuario
        cursor.execute("""
            SELECT
                hr.created_at,
                hr.tipo_sesion,
                hr.posicion,
                hr.puntosmoneda_ganados,
                hr.puntos_totales,
                CASE hr.posicion
                    WHEN 1 THEN '🥇'
                    WHEN 2 THEN '🥈'
                    WHEN 3 THEN '🥉'
                END as medalla
            FROM historial_recompensas hr
            WHERE hr.user_id = %s
            ORDER BY hr.created_at DESC
            LIMIT 10
        """, (g.user['id'],))

        historial = cursor.fetchall()

        # Obtener estadísticas
        cursor.execute("""
            SELECT
                COUNT(*) as total_juegos_premiados,
                SUM(CASE WHEN posicion = 1 THEN 1 ELSE 0 END) as primeros_lugares,
                SUM(CASE WHEN posicion = 2 THEN 1 ELSE 0 END) as segundos_lugares,
                SUM(CASE WHEN posicion = 3 THEN 1 ELSE 0 END) as terceros_lugares,
                SUM(puntosmoneda_ganados) as total_ganado
            FROM historial_recompensas
            WHERE user_id = %s
        """, (g.user['id'],))

        estadisticas = cursor.fetchone()

        # Calcular porcentaje de progreso
        if usuario_info and usuario_info['puntos_maximos'] < 99999999:
            rango_size = usuario_info['puntos_maximos'] - usuario_info['puntos_minimos']
            progreso_actual = usuario_info['puntosmoneda'] - usuario_info['puntos_minimos']
            porcentaje_progreso = (progreso_actual / rango_size * 100) if rango_size > 0 else 0
        else:
            porcentaje_progreso = 100  # Rango máximo

        return render_template('mis_puntos.html',
                             title='Mis Puntos',
                             es_docente=False,
                             usuario_info=usuario_info,
                             todos_rangos=todos_rangos,
                             historial=historial,
                             estadisticas=estadisticas,
                             porcentaje_progreso=porcentaje_progreso)

    except Exception as e:
        app.logger.error(f"Error al cargar mis puntos: {str(e)}")
        flash('Error al cargar la información de puntos', 'error')
        return redirect(url_for('home'))

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@app.route('/update_profile', methods=['POST'])
@jwt_required()
def update_profile():
    if g.user is None and current_identity:
        g.user = current_identity.to_dict()
    if g.user is None:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    user_id = g.user['id']

    try:
        # Actualizar username si se proporcionó
        if 'username' in data:
            success, message = user_controller.actualizar_username(user_id, data['username'])
            if not success:
                return jsonify({'success': False, 'message': message}), 400

        # Actualizar email si se proporcionó
        if 'email' in data:
            success, message = user_controller.actualizar_email(user_id, data['email'])
            if not success:
                return jsonify({'success': False, 'message': message}), 400

        return jsonify({'success': True, 'message': 'Perfil actualizado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/change_password', methods=['POST'])
@jwt_required()
def change_password():
    if g.user is None and current_identity:
        g.user = current_identity.to_dict()
    if g.user is None:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    user_id = g.user['id']
    old_password = data.get('old_password')
    new_password = data.get('new_password')

    if not old_password or not new_password:
        return jsonify({'success': False, 'message': 'Faltan datos'}), 400

    try:
        success, message = user_controller.actualizar_password(user_id, old_password, new_password)
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================
#  Vistas de cuestionarios
# =====================
@app.route('/my-quizzes')
def my_quizzes():
    if g.user is None:
        return redirect(url_for('login'))

    # Solo docentes pueden ver cuestionarios
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden acceder a los cuestionarios', 'error')
        return redirect(url_for('home'))

    db = bd.obtener_conexion()
    created = quiz_controller.listar_cuestionarios_usuario(db, g.user['id'])
    db.close()

    # Por ahora, completados está vacío (se implementará en el futuro)
    completed = []

    return render_template('my_quizzes.html', title='Mis cuestionarios', created=created, completed=completed)

@app.route('/lobby/<int:cuestionario_id>')
def lobby(cuestionario_id):
    """Muestra el lobby de espera para un cuestionario - REHECHA DESDE CERO"""
    if not g.user:
        flash('Debes iniciar sesión para acceder a esta página.', 'warning')
        return redirect(url_for('login'))

    # Solo docentes pueden iniciar sesiones
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden iniciar sesiones de cuestionarios', 'error')
        return redirect(url_for('home'))

    modo = request.args.get('modo', 'individual')

    # Si el modo es grupal, usar EXACTAMENTE la misma lógica que grupos
    if modo == 'grupal':
        try:
            db = bd.obtener_conexion()
            cursor = db.cursor()

            # Verificar que el cuestionario existe
            cursor.execute("SELECT id FROM cuestionarios WHERE id = %s", (cuestionario_id,))
            if not cursor.fetchone():
                flash('Cuestionario no encontrado.', 'error')
                return redirect(url_for('my_quizzes'))

            # Generar session_code único (COPIA EXACTA de api_iniciar_cuestionario_grupo)
            import string
            import random
            while True:
                session_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
                cursor.execute("SELECT id FROM sesiones_grupo WHERE session_code = %s", (session_code,))
                if not cursor.fetchone():
                    break

            # Crear grupo temporal público para esta sala
            import string
            import random
            # Generar código único de grupo
            while True:
                codigo_grupo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
                cursor.execute("SELECT id FROM grupos WHERE codigo = %s", (codigo_grupo,))
                if not cursor.fetchone():
                    break

            nombre_grupo = 'Sala Temporal'
            descripcion = 'Sala temporal creada desde el quiz'
            es_publico = 1

            cursor.execute("""
                INSERT INTO grupos (nombre, descripcion, codigo, es_publico, admin_id, created_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
            """, (nombre_grupo, descripcion, codigo_grupo, es_publico, g.user['id']))

            grupo_id_nuevo = cursor.lastrowid

            # Registrar al creador como miembro del grupo temporal
            cursor.execute("""
                INSERT INTO grupo_miembros (grupo_id, user_id, joined_at)
                VALUES (%s, %s, NOW())
            """, (grupo_id_nuevo, g.user['id']))

            # Crear sesión usando el grupo temporal
            cursor.execute("""
                INSERT INTO sesiones_grupo (grupo_id, cuestionario_id, iniciado_por, created_at, session_code)
                VALUES (%s, %s, %s, NOW(), %s)
            """, (grupo_id_nuevo, cuestionario_id, g.user['id'], session_code))

            sesion_id = cursor.lastrowid
            db.commit()
            cursor.close()
            db.close()

            # Redirigir directamente a grupo_quiz (COPIA EXACTA de grupos)
            return redirect(url_for('grupo_quiz', sesion_id=sesion_id))

        except Exception as e:
            flash(f'Error al crear la sesión: {str(e)}', 'error')
            return redirect(url_for('my_quizzes'))

    # Modo individual (sin cambios)
    db = bd.obtener_conexion()
    cursor = db.cursor()
    cursor.execute("SELECT id, titulo, pin FROM cuestionarios WHERE id = %s", (cuestionario_id,))
    quiz = cursor.fetchone()
    cursor.close()
    db.close()

    if not quiz:
        flash('Cuestionario no encontrado.', 'error')
        return redirect(url_for('my_quizzes'))

    return render_template('lobby.html', quiz=quiz, pin=quiz['pin'], modo=modo)

@app.route('/cuestionario/<pin>/espera')
def espera_cuestionario(pin):
    """Página de espera para un jugador que se une a un cuestionario."""
    db = bd.obtener_conexion()
    cursor = db.cursor()

    # Buscar el cuestionario por PIN
    cursor.execute("SELECT id, titulo FROM cuestionarios WHERE pin = %s", (pin,))
    quiz = cursor.fetchone()

    cursor.close()
    db.close()

    if not quiz:
        flash('El PIN del cuestionario no es válido o ha expirado.', 'error')
        return redirect(url_for('join_quiz'))

    return render_template('lobby.html', quiz=quiz, pin=pin, modo='individual')

@app.route('/api/login', methods=['POST'])
def api_login():
    """
    Endpoint de autenticación JWT para APIs.

    Request Body (JSON):
        {
            "email": "usuario@ejemplo.com",
            "password": "contraseña"
        }

    Response (Success):
        {
            "access_token": "eyJ0eXAiOiJKV1QiLCJhbGc...",
            "user": {
                "id": 123,
                "username": "usuario@ejemplo.com",
                "role": "alumno"
            }
        }

    Response (Error):
        {
            "error": "Descripción del error"
        }
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({'error': 'No se recibieron datos JSON'}), 400

        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            return jsonify({'error': 'Email y password son requeridos'}), 400

        # Autenticar usuario
        user_obj = authenticate(email, password)

        if not user_obj:
            return jsonify({'error': 'Credenciales inválidas'}), 401

        # Generar token JWT
        payload = build_jwt_payload(user_obj)
        token = _encode_jwt_payload(payload)

        return jsonify({
            'access_token': token,
            'user': {
                'id': user_obj.id,
                'username': user_obj.username,
                'role': user_obj.role
            }
        }), 200

    except Exception as e:
        app.logger.error(f"Error en api_login: {str(e)}", exc_info=True)
        return jsonify({
            'error': 'Error interno del servidor',
            'message': str(e) if app.debug else 'Ha ocurrido un error inesperado'
        }), 500


@app.route('/api/test-auth', methods=['POST'])
def api_test_auth():
    """Endpoint de prueba para autenticación JWT con debugging detallado"""
    try:
        data = request.get_json()
        print(f"DEBUG api_test_auth - Datos recibidos: {data}")

        if not data:
            return jsonify({'error': 'No se recibieron datos JSON'}), 400

        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            return jsonify({'error': 'Email y password son requeridos'}), 400

        print(f"DEBUG api_test_auth - Intentando autenticar: {email}")

        # Llamar a la función authenticate
        user_obj = authenticate(email, password)

        if not user_obj:
            print(f"DEBUG api_test_auth - Autenticación fallida")
            return jsonify({'error': 'Credenciales inválidas'}), 401

        print(f"DEBUG api_test_auth - Usuario autenticado: {user_obj.id}")

        # Intentar generar el token
        payload = build_jwt_payload(user_obj)
        print(f"DEBUG api_test_auth - Payload construido: {payload}")

        token = _encode_jwt_payload(payload)
        print(f"DEBUG api_test_auth - Token generado exitosamente")

        return jsonify({
            'access_token': token,
            'user': {
                'id': user_obj.id,
                'username': user_obj.username,
                'role': user_obj.role
            }
        })

    except Exception as e:
        print(f"ERROR en api_test_auth: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': 'Error interno del servidor',
            'message': str(e)
        }), 500

@app.route('/api/verify-token', methods=['POST'])
def api_verify_token():
    """
    Verifica si un token JWT es válido y devuelve información del usuario.

    Request Headers:
        Authorization: JWT <token>

    O Request Body (JSON):
        {
            "token": "eyJ0eXAiOiJKV1QiLCJhbGc..."
        }

    Response (Success):
        {
            "valid": true,
            "user": {
                "id": 123,
                "username": "usuario@ejemplo.com",
                "role": "alumno"
            }
        }

    Response (Error):
        {
            "valid": false,
            "error": "Token inválido o expirado"
        }
    """
    try:
        # Intentar obtener el token del header Authorization
        auth_header = request.headers.get('Authorization')
        token = None

        if auth_header:
            parts = auth_header.split()
            if len(parts) == 2 and parts[0] == 'JWT':
                token = parts[1]

        # Si no está en el header, intentar obtenerlo del body
        if not token:
            data = request.get_json()
            if data:
                token = data.get('token')

        if not token:
            return jsonify({
                'valid': False,
                'error': 'Token no proporcionado'
            }), 400

        # Verificar el token
        payload = _decode_jwt_payload(token)
        user = identity(payload)

        if not user:
            return jsonify({
                'valid': False,
                'error': 'Usuario no encontrado'
            }), 401

        return jsonify({
            'valid': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role
            }
        }), 200

    except ValueError as e:
        return jsonify({
            'valid': False,
            'error': str(e)
        }), 401
    except Exception as e:
        app.logger.error(f"Error en api_verify_token: {str(e)}", exc_info=True)
        return jsonify({
            'valid': False,
            'error': 'Error al verificar el token'
        }), 500


@app.route('/api/refresh-token', methods=['POST'])
def api_refresh_token():
    """
    Refresca un token JWT existente generando uno nuevo.

    Request Headers:
        Authorization: JWT <token>

    Response (Success):
        {
            "access_token": "eyJ0eXAiOiJKV1QiLCJhbGc...",
            "user": {
                "id": 123,
                "username": "usuario@ejemplo.com",
                "role": "alumno"
            }
        }

    Response (Error):
        {
            "error": "Descripción del error"
        }
    """
    try:
        # Obtener el token del header
        auth_header = request.headers.get('Authorization')

        if not auth_header:
            return jsonify({'error': 'Token no proporcionado'}), 400

        parts = auth_header.split()
        if len(parts) != 2 or parts[0] != 'JWT':
            return jsonify({'error': 'Formato de autorización inválido. Use: JWT <token>'}), 400

        token = parts[1]

        # Verificar el token actual (incluso si está expirado, queremos el user_id)
        try:
            payload = _decode_jwt_payload(token)
        except ValueError:
            # Si el token está expirado, intentar decodificarlo sin verificar expiración
            try:
                header_b64, payload_b64, signature_b64 = token.split('.')
                payload = json.loads(_b64url_decode(payload_b64))
            except:
                return jsonify({'error': 'Token inválido'}), 401

        # Obtener el usuario
        user = identity(payload)

        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 401

        # Generar un nuevo token
        new_payload = build_jwt_payload(user)
        new_token = _encode_jwt_payload(new_payload)

        return jsonify({
            'access_token': new_token,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role
            }
        }), 200

    except Exception as e:
        app.logger.error(f"Error en api_refresh_token: {str(e)}", exc_info=True)
        return jsonify({
            'error': 'Error al refrescar el token',
            'message': str(e) if app.debug else 'Ha ocurrido un error inesperado'
        }), 500


@app.route('/api/participantes')
@jwt_required()
def api_participantes():
    """Devuelve la lista de participantes para un PIN de cuestionario."""
    pin = request.args.get('pin')
    if not pin:
        return jsonify({'error': 'PIN no proporcionado'}), 400

    # TODO: Replace with a real database query
    participantes = [
        {'nombre': g.user['username'] if g.user else 'Jugador Anonimo', 'puntaje': 0},
    ]

    return jsonify({'participantes': participantes})

@app.route('/api/iniciar_sesion/<pin>', methods=['POST'])
@jwt_required()
def api_iniciar_sesion(pin):
    """Marca una sesión de cuestionario como iniciada."""
    if not g.user:
        return jsonify({'error': 'No autorizado'}), 401

    # TODO: Update the session status in the database

    return jsonify({'success': True, 'message': 'Cuestionario iniciado'})

@app.route('/sesion/<pin>/pregunta/', defaults={'num_pregunta': 1})
@app.route('/sesion/<pin>/pregunta/<int:num_pregunta>')
def sesion_pregunta(pin, num_pregunta):
    """Muestra una pregunta de un cuestionario en una sesión de juego."""
    if not g.user:
        flash('Debes iniciar sesión para jugar.', 'warning')
        return redirect(url_for('login'))

    db = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener el cuestionario a partir del PIN
        cursor.execute("SELECT id, titulo FROM cuestionarios WHERE pin = %s", (pin,))
        quiz = cursor.fetchone()

        if not quiz:
            cursor.close(); db.close()
            flash('El PIN del cuestionario no es válido o ha expirado.', 'error')
            return redirect(url_for('join_quiz'))

        cuestionario_id = quiz['id']

        # Obtener todas las preguntas con opciones
        cursor.execute("""
            SELECT p.id, p.texto_pregunta, p.tiempo_limite
            FROM preguntas p
            WHERE p.cuestionario_id = %s
            ORDER BY p.orden ASC
        """, (cuestionario_id,))
        preguntas_db = cursor.fetchall()

        preguntas_list = []
        for p in preguntas_db:
            cursor.execute("""
                SELECT id, texto_opcion, es_correcta, orden
                FROM opciones_respuesta
                WHERE pregunta_id = %s
                ORDER BY orden ASC
            """, (p['id'],))
            opciones = cursor.fetchall()
            preguntas_list.append({
                'id': p['id'],
                'texto_pregunta': p['texto_pregunta'],
                'tiempo_limite': p.get('tiempo_limite', 30),
                'opciones': [
                    {'id': o['id'], 'texto_opcion': o['texto_opcion'], 'es_correcta': bool(o['es_correcta'])}
                    for o in opciones
                ]
            })

        cursor.close()

        if not preguntas_list:
            db.close()
            flash('Este cuestionario no tiene preguntas.', 'error')
            return redirect(url_for('my_quizzes'))

        # Si el número de pregunta solicitado es mayor que el total, el quiz ha terminado.
        if num_pregunta > len(preguntas_list):
            db.close()
            return redirect(url_for('resultados_quiz', pin=pin))

        pregunta_actual = preguntas_list[num_pregunta - 1]

        resp = render_template(
            'juego_pregunta.html',
            pin=pin,
            quiz=quiz,
            preguntas=preguntas_list,
            preguntas_json=preguntas_list, # Pasa la lista de Python directamente
            num_pregunta=num_pregunta,
            puntaje_inicial=0,
            tiempo_limite=pregunta_actual['tiempo_limite']
        )
        db.close()
        return resp

    except Exception as e:
        if db:
            try: db.close()
            except: pass
        flash(f'Error al cargar la pregunta: {str(e)}', 'error')
        return redirect(url_for('home'))

@app.route('/api/sesion/responder', methods=['POST'])
@jwt_required()
def api_responder_pregunta():
    """API para que un usuario individual guarde su respuesta."""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    pin = data.get('pin')
    pregunta_id = data.get('pregunta_id')
    opcion_id = data.get('opcion_id')
    tiempo_respuesta = data.get('tiempo_respuesta')

    if not all([pin, pregunta_id, opcion_id, tiempo_respuesta is not None]):
        return jsonify({'success': False, 'message': 'Faltan datos en la solicitud'}), 400

    db = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # 1. Obtener la sesión de juego y el cuestionario
        cursor.execute("""
            SELECT sj.id as sesion_id
            FROM sesiones_juego sj JOIN cuestionarios c ON sj.cuestionario_id = c.id
            WHERE c.pin = %s AND sj.user_id = %s AND sj.estado = 'en_progreso'
            LIMIT 1
        """, (pin, g.user['id']))
        sesion_info = cursor.fetchone()

        if not sesion_info:
            # Si no existe, la creamos
            cursor.execute("SELECT id FROM cuestionarios WHERE pin = %s", (pin,))
            quiz = cursor.fetchone()
            if not quiz:
                return jsonify({'success': False, 'message': 'PIN no válido'}), 404

            # Usamos la nueva columna user_id y la columna created_by para el creador
            cursor.execute("""
                INSERT INTO sesiones_juego (cuestionario_id, user_id, estado, fecha_inicio, created_by)
                VALUES (%s, %s, 'en_progreso', NOW(), %s)
            """, (quiz['id'], g.user['id'], g.user['id']))
            sesion_id = cursor.lastrowid
        else:
            sesion_id = sesion_info['sesion_id']

        # 2. Verificar si la respuesta es correcta y calcular puntos
        cursor.execute("SELECT es_correcta FROM opciones_respuesta WHERE id = %s", (opcion_id,))
        opcion = cursor.fetchone()
        es_correcta = opcion['es_correcta'] if opcion else False

        cursor.execute("SELECT tiempo_limite, puntos FROM preguntas WHERE id = %s", (pregunta_id,))
        pregunta = cursor.fetchone()
        puntos_base = pregunta['puntos'] if pregunta else 1000

        puntos_obtenidos = 0
        if es_correcta:
            # Fórmula de puntos: (1 - (tiempo_respuesta / tiempo_limite / 2)) * puntos_base
            tiempo_limite = pregunta['tiempo_limite'] if pregunta else 30
            puntos_obtenidos = round((1 - (float(tiempo_respuesta) / tiempo_limite / 2)) * puntos_base)

        # 3. Guardar la respuesta
        cursor.execute("""
            INSERT INTO respuestas_participantes (sesion_juego_id, user_id, pregunta_id, opcion_id, es_correcta, puntos_obtenidos, tiempo_respuesta)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (sesion_id, g.user['id'], pregunta_id, opcion_id, es_correcta, puntos_obtenidos, tiempo_respuesta))

        db.commit()
        return jsonify({'success': True, 'correct': bool(es_correcta), 'points': puntos_obtenidos})

    except Exception as e:
        if db: db.rollback()
        return jsonify({'success': False, 'message': f'Error al guardar respuesta: {str(e)}'}), 500

@app.route('/explore')
def explore():
    """Explorar cuestionarios públicos (versión que consulta BD)."""
    if not g.user:
        flash('Debes iniciar sesión para explorar cuestionarios', 'warning')
        return redirect(url_for('login'))

    # Solo docentes pueden explorar cuestionarios
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden explorar cuestionarios', 'error')
        return redirect(url_for('home'))

    search_query = request.args.get('q', '').strip()

    db = bd.obtener_conexion()
    cursor = db.cursor()

    if search_query:
        # Búsqueda por título o PIN
        cursor.execute(
            """
            SELECT id,
                   COALESCE(titulo, 'Sin título') as titulo,
                   pin,
                   COALESCE(descripcion, '') as descripcion,
                   imagen_portada
            FROM cuestionarios
            WHERE estado = 'publico'
            AND (LOWER(titulo) LIKE LOWER(%s) OR CAST(pin AS CHAR) LIKE %s)
            ORDER BY created_at DESC
            """,
            (f'%{search_query}%', f'%{search_query}%')
        )
    else:
        # Sin búsqueda, mostrar todos
        cursor.execute(
            """
            SELECT id,
                   COALESCE(titulo, 'Sin título') as titulo,
                   pin,
                   COALESCE(descripcion, '') as descripcion,
                   imagen_portada
            FROM cuestionarios
            WHERE estado = 'publico'
            ORDER BY created_at DESC
            """
        )

    items = cursor.fetchall()
    cursor.close()
    db.close()
    return render_template('explore.html', title='Explorar', items=items, search_query=search_query)


@app.route('/join')
def join_quiz():
    """Página para unirse a un cuestionario usando un PIN."""
    return render_template('join.html', title='Unirse a Cuestionario')


@app.route('/editor')
def editor():
    # Verificar que solo los docentes puedan acceder
    if g.user is None:
        flash('Debes iniciar sesión para acceder al editor', 'warning')
        return redirect(url_for('login'))

    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden crear cuestionarios', 'error')
        return redirect(url_for('home'))

    return render_template('editor.html', title='Editor', creating_quiz=True)


@app.route('/quiz/<int:cuestionario_id>')
def quiz_details(cuestionario_id):
    """Ver detalles de un cuestionario público o del usuario actual"""
    if not g.user:
        flash('Debes iniciar sesión para ver cuestionarios', 'warning')
        return redirect(url_for('login'))

    # Solo docentes pueden ver detalles de cuestionarios
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden ver cuestionarios', 'error')
        return redirect(url_for('home'))

    db = bd.obtener_conexion()
    cursor = db.cursor()

    # Obtener cuestionario
    cursor.execute("""
        SELECT c.id, c.user_id, c.titulo, c.descripcion, c.imagen_portada, c.pin, c.estado, c.created_at,
               u.username as creator_username
        FROM cuestionarios c
        LEFT JOIN users u ON c.user_id = u.id
        WHERE c.id = %s
    """, (cuestionario_id,))

    cuestionario = cursor.fetchone()

    if not cuestionario:
        cursor.close()
        db.close()
        flash('Cuestionario no encontrado', 'error')
        return redirect(url_for('explore'))

    # Verificar permisos:
    # - Si es público, todos pueden verlo
    # - Si es privado, solo el creador puede verlo en esta vista
    if cuestionario['estado'] == 'privado':
        if not g.user or g.user['id'] != cuestionario['user_id']:
            cursor.close()
            db.close()
            flash('Este cuestionario es privado. Usa el PIN para acceder.', 'warning')
            return redirect(url_for('join_quiz'))

    # Obtener preguntas
    cursor.execute("""
        SELECT id, tipo_pregunta, texto_pregunta, orden, tiempo_limite, puntos
        FROM preguntas
        WHERE cuestionario_id = %s
        ORDER BY orden ASC
    """, (cuestionario_id,))

    preguntas = cursor.fetchall()

    # Obtener opciones para cada pregunta
    preguntas_con_opciones = []
    for pregunta in preguntas:
        cursor.execute("""
            SELECT id, texto_opcion, es_correcta, orden
            FROM opciones_respuesta
            WHERE pregunta_id = %s
            ORDER BY orden ASC
        """, (pregunta['id'],))

        opciones = cursor.fetchall()

        preguntas_con_opciones.append({
            'id': pregunta['id'],
            'texto': pregunta['texto_pregunta'],
            'answers': [{'texto': op['texto_opcion'], 'is_correct': bool(op['es_correcta'])} for op in opciones]
        })

    cursor.close()
    db.close()

    # Construir URL completa de la imagen
    if cuestionario['imagen_portada'] and cuestionario['imagen_portada'].strip():
        image_url = url_for('static', filename='uploads/' + cuestionario['imagen_portada'])
    else:
        image_url = url_for('static', filename='img/sinimagenes.jpeg')

    quiz_data = {
        'id': cuestionario['id'],
        'titulo': cuestionario['titulo'],
        'descripcion': cuestionario['descripcion'],
        'image_url': image_url,
        'pin': cuestionario['pin'],
        'creator_username': cuestionario['creator_username'],
        'questions': preguntas_con_opciones
    }

    return render_template('quiz_details.html', title=quiz_data['titulo'], quiz=quiz_data)


@app.route('/editor/<int:cuestionario_id>')
def editor_edit(cuestionario_id):
    """Editar un cuestionario existente"""
    if g.user is None:
        flash('Debes iniciar sesión para editar cuestionarios', 'warning')
        return redirect(url_for('login'))

    # Solo docentes pueden editar cuestionarios
    if g.user.get('role') != 'docente':
        flash('Solo los docentes pueden editar cuestionarios', 'error')
        return redirect(url_for('home'))

    db = bd.obtener_conexion()
    response = quiz_controller.obtener_cuestionario(db, cuestionario_id)
    db.close()

    # Si hay error, redirigir a my_quizzes
    if response[1] != 200:
        flash('Cuestionario no encontrado', 'error')
        return redirect(url_for('my_quizzes'))

    cuestionario = response[0].get_json()['cuestionario']
    return render_template('editor.html', title='Editor', creating_quiz=False, cuestionario=cuestionario)


# =====================
#  API de cuestionarios
# =====================
@app.route('/api/cuestionario', methods=['POST'])
@jwt_required()
def crear_cuestionario():
    """Crear un nuevo cuestionario"""
    if g.user is None:
        return jsonify({'error': 'No autorizado'}), 401

    # Verificar que solo los docentes puedan crear cuestionarios
    if g.user.get('role') != 'docente':
        return jsonify({'error': 'Solo los docentes pueden crear cuestionarios'}), 403

    try:
        data = {}

        if request.is_json:
            json_data = request.get_json()
            data['titulo'] = json_data.get('titulo')
            data['descripcion'] = json_data.get('descripcion')
            data['preguntas'] = json_data.get('preguntas', [])
            data['pin'] = json_data.get('pin', '')
            data['estado'] = json_data.get('estado', 'publico')
        else:
            data['titulo'] = request.form.get('titulo')
            data['descripcion'] = request.form.get('descripcion')
            data['pin'] = request.form.get('pin', '')
            data['estado'] = request.form.get('estado', 'publico')

            import json
            preguntas_str = request.form.get('preguntas', '[]')
            try:
                data['preguntas'] = json.loads(preguntas_str) if preguntas_str else []
            except Exception:
                data['preguntas'] = []

        db = bd.obtener_conexion()
        response = quiz_controller.crear_cuestionario(db, data, request.files, app.config['UPLOAD_FOLDER'])
        db.close()
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cuestionario/<int:cuestionario_id>', methods=['PUT', 'POST'])
@jwt_required()
def actualizar_cuestionario(cuestionario_id):
    """Actualizar un cuestionario existente"""
    if g.user is None:
        return jsonify({'error': 'No autorizado'}), 401

    # Verificar que solo los docentes puedan actualizar cuestionarios
    if g.user.get('role') != 'docente':
        return jsonify({'error': 'Solo los docentes pueden actualizar cuestionarios'}), 403

    try:
        data = {}

        if request.is_json:
            json_data = request.get_json()
            data['titulo'] = json_data.get('titulo')
            data['descripcion'] = json_data.get('descripcion')
            data['preguntas'] = json_data.get('preguntas', [])
            data['pin'] = json_data.get('pin', '')
            data['estado'] = json_data.get('estado', 'publico')
        else:
            data['titulo'] = request.form.get('titulo')
            data['descripcion'] = request.form.get('descripcion')
            data['pin'] = request.form.get('pin', '')
            data['estado'] = request.form.get('estado', 'publico')

            import json
            preguntas_str = request.form.get('preguntas', '[]')
            try:
                data['preguntas'] = json.loads(preguntas_str) if preguntas_str else []
            except Exception:
                data['preguntas'] = []

        db = bd.obtener_conexion()
        response = quiz_controller.actualizar_cuestionario(
            db, cuestionario_id, data, request.files, app.config['UPLOAD_FOLDER']
        )
        db.close()
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cuestionario/<int:cuestionario_id>', methods=['GET'])
@jwt_required()
def obtener_cuestionario(cuestionario_id):
    """Obtener un cuestionario con todas sus preguntas"""
    if g.user is None:
        return jsonify({'error': 'No autorizado'}), 401

    # Solo docentes pueden obtener cuestionarios
    if g.user.get('role') != 'docente':
        return jsonify({'error': 'Solo los docentes pueden obtener cuestionarios'}), 403

    db = bd.obtener_conexion()
    response = quiz_controller.obtener_cuestionario(db, cuestionario_id)
    db.close()
    return response


@app.route('/api/cuestionario/<int:cuestionario_id>', methods=['DELETE'])
@jwt_required()
def eliminar_cuestionario(cuestionario_id):
    """Eliminar un cuestionario"""
    if g.user is None:
        return jsonify({'error': 'No autorizado'}), 401

    # Solo docentes pueden eliminar cuestionarios
    if g.user.get('role') != 'docente':
        return jsonify({'error': 'Solo los docentes pueden eliminar cuestionarios'}), 403

    db = bd.obtener_conexion()
    response = quiz_controller.eliminar_cuestionario(db, cuestionario_id)
    db.close()
    return response


@app.route('/api/importar-preguntas-excel', methods=['POST'])
@jwt_required()
def importar_preguntas_excel():
    """Importar preguntas desde un archivo Excel"""
    if g.user is None:
        return jsonify({'error': 'No autorizado'}), 401

    # Solo docentes pueden importar preguntas
    if g.user.get('role') != 'docente':
        return jsonify({'error': 'Solo los docentes pueden importar preguntas'}), 403

    try:
        # Verificar que se haya enviado un archivo
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No se encontró el archivo Excel'}), 400

        file = request.files['excel_file']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400

        # Validar extensión
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'El archivo debe ser un archivo Excel (.xlsx o .xls)'}), 400

        # Leer el archivo Excel
        from openpyxl import load_workbook

        try:
            # Cargar el workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            preguntas = []

            # Saltar la fila de encabezados (fila 1)
            for row_num in range(2, ws.max_row + 1):
                # Leer datos de la fila
                pregunta_texto = ws.cell(row=row_num, column=1).value
                tipo_pregunta = ws.cell(row=row_num, column=2).value
                puntos = ws.cell(row=row_num, column=3).value
                tiempo = ws.cell(row=row_num, column=4).value

                # Leer respuestas (columnas 5-9: respuesta 1-5)
                respuestas_texto = []
                for col in range(5, 10):
                    resp = ws.cell(row=row_num, column=col).value
                    if resp:
                        respuestas_texto.append(str(resp).strip())

                # Leer respuestas correctas (columnas 10-14: correcta 1-5)
                respuestas_correctas = []
                for col in range(10, 15):
                    correcta = ws.cell(row=row_num, column=col).value
                    # Considerar como correcta si es TRUE, 1, "SI", "X", etc.
                    if correcta:
                        correcta_str = str(correcta).strip().upper()
                        es_correcta = correcta_str in ['TRUE', '1', 'SI', 'SÍ', 'SÍ', 'X', 'YES', 'V', 'VERDADERO']
                        respuestas_correctas.append(es_correcta)
                    else:
                        respuestas_correctas.append(False)

                # Validaciones básicas
                if not pregunta_texto or str(pregunta_texto).strip() == '':
                    continue  # Saltar filas vacías

                # Validar tipo de pregunta
                tipo_pregunta_str = str(tipo_pregunta).strip().lower() if tipo_pregunta else 'multiple'
                if tipo_pregunta_str not in ['multiple', 'simple', 'verdadero-falso']:
                    tipo_pregunta_str = 'multiple'

                # Validar puntos
                try:
                    puntos = int(puntos) if puntos else 1
                    if puntos < 1 or puntos > 5:
                        puntos = 1
                except:
                    puntos = 1

                # Validar tiempo
                try:
                    tiempo = int(tiempo) if tiempo else 30
                    if tiempo not in [10, 20, 30, 45, 60, 90]:
                        tiempo = 30
                except:
                    tiempo = 30

                # Construir las respuestas
                respuestas = []
                for i, texto in enumerate(respuestas_texto):
                    if texto:  # Solo agregar respuestas no vacías
                        es_correcta = respuestas_correctas[i] if i < len(respuestas_correctas) else False
                        respuestas.append({
                            'text': texto,
                            'isCorrect': es_correcta
                        })

                # Validar que haya al menos 2 respuestas
                if len(respuestas) < 2:
                    continue

                # Validar que haya al menos una respuesta correcta
                if not any(r['isCorrect'] for r in respuestas):
                    continue

                # Agregar la pregunta
                preguntas.append({
                    'text': str(pregunta_texto).strip(),
                    'type': tipo_pregunta_str,
                    'points': puntos,
                    'time': tiempo,
                    'answers': respuestas
                })

            if len(preguntas) == 0:
                return jsonify({
                    'success': False,
                    'error': 'No se encontraron preguntas válidas en el archivo. Verifica el formato.'
                }), 400

            return jsonify({
                'success': True,
                'preguntas': preguntas,
                'message': f'Se importaron {len(preguntas)} pregunta(s) correctamente'
            })

        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error al procesar el archivo Excel: {str(e)}'
            }), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/Contactanos')
def Contactanos():
    return render_template('contact.html')

@app.route('/Nosotros')
def Nosotros():
    return render_template('about_us.html')

@app.route('/Desarrolladores')
def Desarrolladores():
    return render_template('developers.html')

@app.route('/Preguntas_Frecuentes')
def Preguntas_Frecuentes():
    return render_template('faq.html')

@app.route('/Soporte_Tecnico')
def Soporte_Tecnico():
    return render_template('technical_support.html')
# =====================
#  Main
# =====================

@app.route('/delete_account', methods=['POST'])
def delete_account():
    """Eliminar la cuenta del usuario actual"""
    if g.user is None:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    user_id = g.user['id']

    try:
        db = bd.obtener_conexion()
        success, message = user_controller.eliminar_usuario(db, user_id)
        db.close()

        if success:
            session.clear()  # cerrar sesión tras eliminar cuenta
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        print("Error al eliminar cuenta:", e)
        return jsonify({'success': False, 'message': 'Error interno al eliminar cuenta'}), 500

@app.route('/recuperar_cuenta')
def recuperar_cuenta():
    return render_template('recuperar_cuenta.html')

@app.route('/verificar_recuperar_cuenta')
def verificar_recuperar_cuenta():
    return render_template('verificar_recuperar_cuenta.html')

@app.route('/cambiar_contrasena', methods=['GET', 'POST'])
def cambiar_contrasena():
    if request.method == 'POST':
        nueva = request.form['new-password']
        confirmar = request.form['confirm-password']
        email = session.get('email_verificacion')  # el correo que guardaste tras verificar el código

        if nueva != confirmar:
            return render_template('cambiar_contra.html', mensaje="Las contraseñas no coinciden")

        if email:
            user_controller.actualizar_password_por_email(email, nueva)
            session.pop('email_verificacion', None)  # limpia la sesión
            return redirect(url_for('login'))  # redirige al inicio o login

        return render_template('cambiar_contra.html', mensaje="Sesión inválida o expirada")

    return render_template('cambiar_contra.html')

# === RUTAS PARA GRUPOS ===

@app.route('/grupos')
def grupos():
    """Página principal de gestión de grupos"""
    if not g.user:
        flash('Debes iniciar sesión para acceder a los grupos', 'warning')
        return redirect(url_for('login'))

    # Obtener grupos del usuario
    db = bd.obtener_conexion()
    cursor = db.cursor()

    # Grupos donde el usuario es miembro
    cursor.execute("""
        SELECT g.id, g.nombre, g.descripcion, g.codigo, g.es_publico, g.created_at,
               CASE WHEN g.admin_id = %s THEN true ELSE false END as es_admin,
               COUNT(gm.user_id) as miembros_count
        FROM grupos g
        LEFT JOIN grupo_miembros gm ON g.id = gm.grupo_id
        WHERE g.id IN (
            SELECT grupo_id FROM grupo_miembros WHERE user_id = %s
        )
        GROUP BY g.id, g.nombre, g.descripcion, g.codigo, g.es_publico, g.created_at, g.admin_id
        ORDER BY g.created_at DESC
    """, (g.user['id'], g.user['id']))

    mis_grupos = cursor.fetchall()

    # Grupos públicos disponibles (donde el usuario NO es miembro)
    cursor.execute("""
        SELECT g.id, g.nombre, g.descripcion, g.codigo, g.es_publico, g.created_at,
               false as es_admin,
               COUNT(gm.user_id) as miembros_count
        FROM grupos g
        LEFT JOIN grupo_miembros gm ON g.id = gm.grupo_id
        WHERE g.es_publico = true
        AND g.id NOT IN (
            SELECT grupo_id FROM grupo_miembros WHERE user_id = %s
        )
        GROUP BY g.id, g.nombre, g.descripcion, g.codigo, g.es_publico, g.created_at
        ORDER BY g.created_at DESC
    """, (g.user['id'],))

    grupos_publicos = cursor.fetchall()
    cursor.close()
    db.close()

    return render_template('grupos.html',
                         title='Grupos',
                         grupos=mis_grupos,
                         grupos_publicos=grupos_publicos)

@app.route('/api/grupos/crear', methods=['POST'])
@jwt_required()
def api_crear_grupo():
    """API para crear un nuevo grupo"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    # Verificar que solo los docentes puedan crear grupos/salas
    if g.user.get('role') != 'docente':
        return jsonify({'success': False, 'message': 'Solo los docentes pueden crear salas'}), 403

    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    descripcion = data.get('descripcion', '').strip()
    es_publico = data.get('es_publico', False)

    if not nombre:
        return jsonify({'success': False, 'message': 'El nombre del grupo es requerido'})

    # Generar código único de 8 caracteres
    import string
    import random
    codigo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Crear grupo
        cursor.execute("""
            INSERT INTO grupos (nombre, descripcion, codigo, es_publico, admin_id, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """, (nombre, descripcion, codigo, es_publico, g.user['id']))

        grupo_id = cursor.lastrowid

        # Agregar admin como miembro
        cursor.execute("""
            INSERT INTO grupo_miembros (grupo_id, user_id, joined_at)
            VALUES (%s, %s, NOW())
        """, (grupo_id, g.user['id']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'message': 'Grupo creado exitosamente',
            'grupo_id': grupo_id,
            'codigo': codigo
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al crear grupo: {str(e)}'})

@app.route('/api/grupos/unirse', methods=['POST'])
@jwt_required()
def api_unirse_grupo():
    """API para unirse a un grupo por código"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    codigo = data.get('codigo', '').strip().upper()

    if len(codigo) != 6:
        return jsonify({'success': False, 'message': 'El código debe tener 8 caracteres'})

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Buscar grupo por código
        cursor.execute("SELECT id, nombre FROM grupos WHERE codigo = %s", (codigo,))
        grupo = cursor.fetchone()

        if not grupo:
            return jsonify({'success': False, 'message': 'Grupo no encontrado'})

        # Verificar si ya es miembro
        cursor.execute("""
            SELECT id FROM grupo_miembros
            WHERE grupo_id = %s AND user_id = %s
        """, (grupo['id'], g.user['id']))

        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'Ya eres miembro de este grupo'})

        # Agregar como miembro
        cursor.execute("""
            INSERT INTO grupo_miembros (grupo_id, user_id, joined_at)
            VALUES (%s, %s, NOW())
        """, (grupo['id'], g.user['id']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'message': f'Te has unido al grupo "{grupo["nombre"]}"'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al unirse al grupo: {str(e)}'})

@app.route('/api/grupos/salir', methods=['POST'])
@jwt_required()
def api_salir_grupo():
    """API para salir de un grupo"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    grupo_id = data.get('grupo_id')

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar si es miembro
        cursor.execute("""
            SELECT id FROM grupo_miembros
            WHERE grupo_id = %s AND user_id = %s
        """, (grupo_id, g.user['id']))

        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'No eres miembro de este grupo'})

        # Remover del grupo
        cursor.execute("""
            DELETE FROM grupo_miembros
            WHERE grupo_id = %s AND user_id = %s
        """, (grupo_id, g.user['id']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({'success': True, 'message': 'Has salido del grupo'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al salir del grupo: {str(e)}'})

@app.route('/api/grupos/cuestionarios')
@jwt_required()
def api_grupos_cuestionarios():
    """API para obtener cuestionarios disponibles para grupos"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    # Solo docentes pueden obtener cuestionarios
    if g.user.get('role') != 'docente':
        return jsonify({'success': False, 'message': 'Solo los docentes pueden acceder a cuestionarios'}), 403

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener cuestionarios públicos y del usuario
        cursor.execute("""
            SELECT c.id, c.titulo, c.descripcion, c.pin, c.imagen_portada,
                   COUNT(p.id) as preguntas_count
            FROM cuestionarios c
            LEFT JOIN preguntas p ON c.id = p.cuestionario_id
            WHERE c.estado = 'publico' OR c.user_id = %s
            GROUP BY c.id, c.titulo, c.descripcion, c.pin, c.imagen_portada
            ORDER BY c.created_at DESC
        """, (g.user['id'],))

        cuestionarios = cursor.fetchall()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'quizzes': cuestionarios
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al cargar cuestionarios: {str(e)}'})

@app.route('/api/grupos/iniciar-cuestionario', methods=['POST'])
@jwt_required()
def api_iniciar_cuestionario_grupo():
    """API para iniciar un cuestionario en grupo"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    # Solo docentes pueden iniciar cuestionarios en grupo
    if g.user.get('role') != 'docente':
        return jsonify({'success': False, 'message': 'Solo los docentes pueden iniciar cuestionarios'}), 403

    data = request.get_json()
    grupo_id = data.get('grupo_id')
    cuestionario_id = data.get('cuestionario_id')

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar que el usuario es miembro del grupo
        cursor.execute("""
            SELECT id FROM grupo_miembros
            WHERE grupo_id = %s AND user_id = %s
        """, (grupo_id, g.user['id']))

        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'No eres miembro de este grupo'})

        # Generar un session_code único
        import string
        import random
        while True:
            session_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            cursor.execute("SELECT id FROM sesiones_grupo WHERE session_code = %s", (session_code,))
            if not cursor.fetchone():
                break

        # Crear sesión de cuestionario en grupo
        cursor.execute("""
            INSERT INTO sesiones_grupo (grupo_id, cuestionario_id, iniciado_por, created_at, session_code)
            VALUES (%s, %s, %s, NOW(), %s)
        """, (grupo_id, cuestionario_id, g.user['id'], session_code))

        sesion_id = cursor.lastrowid
        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'sesion_id': sesion_id,
            'message': 'Cuestionario iniciado en el grupo'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al iniciar cuestionario: {str(e)}'})

@app.route('/grupo/quiz/<int:sesion_id>')
def grupo_quiz(sesion_id):
    """Página para rendir cuestionario en grupo - REHECHA DESDE CERO COPIANDO GRUPOS"""
    if not g.user:
        flash('Debes iniciar sesión para participar', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión (COPIA EXACTA de grupos)
        cursor.execute("""
            SELECT sg.id, sg.grupo_id, sg.cuestionario_id, sg.estado, sg.session_code, sg.iniciado_por,
                   COALESCE(g.nombre, 'Sala Temporal') as grupo_nombre, COALESCE(g.es_publico, 1) as es_publico,
                   c.titulo, c.pin, c.descripcion
            FROM sesiones_grupo sg
            LEFT JOIN grupos g ON sg.grupo_id = g.id
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            WHERE sg.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('grupos'))

        # Verificar membresía SOLO si hay grupo específico y NO es público
        if sesion['grupo_id'] and not sesion['es_publico']:
            cursor.execute("""
                SELECT id FROM grupo_miembros
                WHERE grupo_id = %s AND user_id = %s
            """, (sesion['grupo_id'], g.user['id']))

            if not cursor.fetchone():
                flash('No eres miembro de este grupo', 'error')
                return redirect(url_for('grupos'))

        # Registrar al usuario en la sesión SOLO si NO es el creador (COPIA EXACTA de grupos)
        if g.user['id'] != sesion['iniciado_por']:
            cursor.execute("""
                INSERT INTO usuario_estado_grupo (sesion_id, user_id, esta_listo)
                VALUES (%s, %s, 0)
                ON DUPLICATE KEY UPDATE user_id = user_id
            """, (sesion_id, g.user['id']))

        db.commit()

        # Obtener miembros (COPIA EXACTA de grupos)
        cursor.execute("""
            SELECT u.id, u.username, ues.esta_listo
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            JOIN sesiones_grupo sg ON ues.sesion_id = sg.id
            WHERE ues.sesion_id = %s AND ues.user_id != sg.iniciado_por
        """, (sesion_id,))
        miembros = cursor.fetchall()

        # Obtener preguntas (COPIA EXACTA de grupos)
        cursor.execute("""
            SELECT id, texto_pregunta, tiempo_limite
            FROM preguntas
            WHERE cuestionario_id = %s
            ORDER BY orden
        """, (sesion['cuestionario_id'],))
        preguntas = cursor.fetchall()

        preguntas_con_opciones = []
        for pregunta in preguntas:
            cursor.execute("""
                SELECT id, texto_opcion, es_correcta
                FROM opciones_respuesta
                WHERE pregunta_id = %s
                ORDER BY orden
            """, (pregunta['id'],))
            opciones = cursor.fetchall()

            preguntas_con_opciones.append({
                'id': pregunta['id'],
                'text': pregunta['texto_pregunta'],
                'time_limit': pregunta['tiempo_limite'],
                'options': [{'id': o['id'], 'text': o['texto_opcion'], 'is_correct': bool(o['es_correcta'])} for o in opciones]
            })

        cursor.close()
        db.close()

        # Verificar si es creador (normaliza tipos a int)
        es_creador = (int(g.user['id']) == int(sesion['iniciado_por']))

        return render_template('grupo_quiz.html',
                             sesion_id=sesion_id,
                             session_code=sesion['session_code'],
                             grupo={'nombre': sesion['grupo_nombre']},
                             cuestionario={
                                 'titulo': sesion['titulo'],
                                 'pin': sesion['pin'],
                                 'preguntas_count': len(preguntas)
                             },
                             miembros=miembros,
                             preguntas_json=json.dumps(preguntas_con_opciones),
                             es_creador=es_creador)

    except Exception as e:
        flash(f'Error al cargar el cuestionario: {str(e)}', 'error')
        return redirect(url_for('grupos'))

@app.route('/unirse-sala')
def unirse_sala():
    """Página para unirse a una sala temporal usando código"""
    if not g.user:
        flash('Debes iniciar sesión para unirte a una sala', 'warning')
        return redirect(url_for('login'))

    return render_template('unirse_sala.html')


@app.route('/api/grupo/unirse-sesion', methods=['POST'])
@jwt_required()
def api_unirse_sesion_grupo():
    """API para unirse a una sesión de grupo usando el código de sala"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    session_code = data.get('session_code', '').strip().upper()

    if not session_code or len(session_code) != 6:
        return jsonify({'success': False, 'message': 'Código de sala inválido'}), 400

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # 1. Buscar la sesión por código (puede ser con o sin grupo específico)
        cursor.execute("""
            SELECT sg.id, sg.grupo_id, sg.estado, sg.session_code,
                   COALESCE(g.nombre, 'Sala Temporal') as grupo_nombre,
                   COALESCE(g.es_publico, 1) as es_publico,
                   c.titulo
            FROM sesiones_grupo sg
            LEFT JOIN grupos g ON sg.grupo_id = g.id
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            WHERE sg.session_code = %s
        """, (session_code,))

        sesion = cursor.fetchone()

        if not sesion:
            cursor.close(); db.close()
            return jsonify({'success': False, 'message': 'Código de sala no encontrado'}), 404

        # 2. Verificar membresía solo si el grupo existe y NO es público (salas temporales públicas permiten acceso libre)
        if sesion['grupo_id'] and not sesion['es_publico']:
            cursor.execute("""
                SELECT id FROM grupo_miembros
                WHERE grupo_id = %s AND user_id = %s
            """, (sesion['grupo_id'], g.user['id']))

            if not cursor.fetchone():
                cursor.close(); db.close()
                return jsonify({'success': False, 'message': 'No eres miembro de este grupo'}), 403
        # Para sesiones sin grupo específico o grupos públicos, cualquier usuario puede unirse

        # 3. Verificar el estado de la sesión
        if sesion['estado'] == 'finalizado':
            cursor.close(); db.close()
            return jsonify({'success': False, 'message': 'Esta sesión ya ha finalizado'}), 400

        # 4. Registrar al usuario en la sesión si no está ya
        cursor.execute("""
            INSERT INTO usuario_estado_grupo (sesion_id, user_id, esta_listo)
            VALUES (%s, %s, 0)
            ON DUPLICATE KEY UPDATE user_id = user_id
        """, (sesion['id'], g.user['id']))

        db.commit()
        cursor.close(); db.close()

        return jsonify({
            'success': True,
            'sesion_id': sesion['id'],
            'grupo_nombre': sesion['grupo_nombre'],
            'message': f'Te has unido a la sala del grupo {sesion["grupo_nombre"]}'
        })

    except Exception as e:
        if db:
            try: db.close()
            except: pass
        return jsonify({'success': False, 'message': f'Error al unirse a la sesión: {str(e)}'}), 500

@app.route('/api/grupo/ready', methods=['POST'])
@jwt_required()
def api_grupo_ready():
    """API para marcar usuario como listo y, si todos los que están en la sala están listos, iniciar la sesión."""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')
    ready = bool(data.get('ready', False))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # 1) Upsert del estado del usuario en la sesión
        cursor.execute("""
            INSERT INTO usuario_estado_grupo (sesion_id, user_id, esta_listo)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE esta_listo = VALUES(esta_listo)
        """, (sesion_id, g.user['id'], ready))

        # 2) Obtener grupo de la sesión
        cursor.execute("SELECT grupo_id FROM sesiones_grupo WHERE id = %s", (sesion_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close(); db.close()
            return jsonify({'success': False, 'message': 'Sesión no encontrada'}), 404
        grupo_id = row['grupo_id']

        # 3) CAMBIO CLAVE: Contar usuarios que se UNIERON A LA SALA (están en usuario_estado_grupo)
        # en lugar de contar todos los miembros del grupo
        cursor.execute("""
            SELECT COUNT(*) AS total_en_sala
            FROM usuario_estado_grupo
            WHERE sesion_id = %s
        """, (sesion_id,))
        total_en_sala = cursor.fetchone()['total_en_sala']

        # 4) Contar cuántos están listos en esta sesión
        cursor.execute("""
            SELECT COUNT(*) AS listos
            FROM usuario_estado_grupo
            WHERE sesion_id = %s AND esta_listo = 1
        """, (sesion_id,))
        listos = cursor.fetchone()['listos']

        # 5) Todos están listos si hay al menos 1 persona en la sala y todos están listos
        all_ready = (total_en_sala > 0 and listos == total_en_sala)

        # 6) Si todos listos, pasar a 'en_progreso' y setear started_at
        if all_ready:
            cursor.execute("""
                UPDATE sesiones_grupo
                SET estado = 'en_progreso', started_at = NOW()
                WHERE id = %s
            """, (sesion_id,))

        # 7) Obtener la lista actualizada SOLO de usuarios que están en la sala
        cursor.execute("""
            SELECT u.id, u.username, ues.esta_listo as ready
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            WHERE ues.sesion_id = %s
        """, (sesion_id,))

        miembros_actualizados = cursor.fetchall()

        db.commit()
        cursor.close(); db.close()

        return jsonify({'success': True, 'all_ready': all_ready, 'members': miembros_actualizados})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500



@app.route('/grupo/juego/<int:sesion_id>')
def grupo_juego(sesion_id):
    """Página para jugar el cuestionario en grupo (solo cuando está en progreso)"""
    if not g.user:
        flash('Debes iniciar sesión para participar', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión
        cursor.execute("""
            SELECT sg.id, sg.grupo_id, sg.cuestionario_id, sg.estado, sg.session_code, sg.iniciado_por,
                   COALESCE(g.nombre, 'Sala Temporal') as grupo_nombre, COALESCE(g.es_publico, 1) as es_publico,
                   c.titulo, c.pin
            FROM sesiones_grupo sg
            LEFT JOIN grupos g ON sg.grupo_id = g.id
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            WHERE sg.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('grupos'))

        # Verificar que el usuario es miembro del grupo (solo si NO es público)
        if sesion['grupo_id'] and not sesion['es_publico']:
            cursor.execute("""
                SELECT id FROM grupo_miembros
                WHERE grupo_id = %s AND user_id = %s
            """, (sesion['grupo_id'], g.user['id']))

            if not cursor.fetchone():
                flash('No eres miembro de este grupo', 'error')
                return redirect(url_for('grupos'))
        else:
            # Para salas temporales, cualquier usuario puede acceder
            pass

        # Verificar estado de registro del usuario en la sesión
        cursor.execute("SELECT iniciado_por FROM sesiones_grupo WHERE id = %s", (sesion_id,))
        fila = cursor.fetchone()
        iniciado_por = fila['iniciado_por'] if fila else None

        cursor.execute("""
            SELECT esta_listo
            FROM usuario_estado_grupo
            WHERE sesion_id = %s AND user_id = %s
        """, (sesion_id, g.user['id']))
        registro = cursor.fetchone()

        # Si es el creador, permitir el acceso aunque no esté registrado
        if int(g.user['id']) != int(iniciado_por):
            # No es creador: debe estar registrado
            if not registro:
                flash('No estás registrado en esta sesión', 'error')
                cursor.close(); db.close()
                return redirect(url_for('grupo_quiz', sesion_id=sesion_id))
            # Ya no se requiere estar "listo" - solo estar registrado

        # Verificar estado de la sesión
        if sesion['estado'] != 'en_progreso':
            if sesion['estado'] == 'esperando':
                return redirect(url_for('grupo_quiz', sesion_id=sesion_id))
            elif sesion['estado'] == 'finalizado':
                return redirect(url_for('grupo_resultados', sesion_id=sesion_id))

        # Obtener preguntas del cuestionario
        cursor.execute("""
            SELECT p.id, p.texto_pregunta, p.tiempo_limite, p.puntos, p.orden
            FROM preguntas p
            WHERE p.cuestionario_id = %s
            ORDER BY p.orden ASC
        """, (sesion['cuestionario_id'],))

        preguntas = cursor.fetchall()

        # Obtener opciones para cada pregunta
        preguntas_con_opciones = []
        for pregunta in preguntas:
            cursor.execute("""
                SELECT id, texto_opcion, es_correcta, orden
                FROM opciones_respuesta
                WHERE pregunta_id = %s
                ORDER BY orden ASC
            """, (pregunta['id'],))
            opciones = cursor.fetchall()
            preguntas_con_opciones.append({
                'id': pregunta['id'],
                'text': pregunta['texto_pregunta'],
                'time_limit': pregunta['tiempo_limite'],
                'points': pregunta['puntos'],
                'orden': pregunta['orden'],
                'options': [{'id': o['id'], 'text': o['texto_opcion'], 'is_correct': bool(o['es_correcta'])} for o in opciones]
            })

        # Obtener miembros en la sesión (EXCLUYENDO al creador)
        cursor.execute("""
            SELECT u.id, u.username
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            JOIN sesiones_grupo sg ON ues.sesion_id = sg.id
            WHERE ues.sesion_id = %s AND ues.user_id != sg.iniciado_por
        """, (sesion_id,))

        miembros = cursor.fetchall()

        cursor.close()
        db.close()

        # Verificar si es creador (normaliza tipos a int)
        es_creador = (int(g.user['id']) == int(sesion['iniciado_por']))

        return render_template('juego_grupo.html',
                             sesion_id=sesion_id,
                             grupo={'nombre': sesion['grupo_nombre']},
                             cuestionario={'titulo': sesion['titulo']},
                             preguntas_json=json.dumps(preguntas_con_opciones),
                             total_preguntas=len(preguntas),
                             miembros=miembros,
                             es_creador=es_creador)

    except Exception as e:
        flash(f'Error al cargar el juego: {str(e)}', 'error')
        return redirect(url_for('grupos'))

@app.route('/api/grupo/miembros/<int:sesion_id>')
@jwt_required()
def api_grupo_miembros(sesion_id):
    """API para obtener la lista actualizada de miembros en una sesión"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener miembros que están en la sesión
        cursor.execute("""
            SELECT u.id, u.username, ues.esta_listo as ready
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            WHERE ues.sesion_id = %s
            ORDER BY ues.created_at ASC
        """, (sesion_id,))

        miembros = cursor.fetchall()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'members': miembros
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/grupo/status/<int:sesion_id>')
@jwt_required()
def api_grupo_status(sesion_id):
    """API para obtener el estado de la sesión"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        cursor.execute("SELECT estado FROM sesiones_grupo WHERE id = %s", (sesion_id,))
        sesion = cursor.fetchone()

        if not sesion:
            return jsonify({'success': False, 'message': 'Sesión no encontrada'})

        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'status': sesion['estado']
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/grupo/answer', methods=['POST'])
@jwt_required()
def api_grupo_answer():
    """API para enviar respuesta del usuario con cálculo de puntos"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')
    question_id = data.get('question_id')
    answer_id = data.get('answer_id')  # Puede ser None para timeout
    tiempo_respuesta = data.get('tiempo_respuesta', 0)

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Si answer_id es None, es un timeout (sin respuesta)
        if answer_id is None or answer_id == 'null':
            # Timeout: registrar sin respuesta
            cursor.execute("""
                INSERT INTO respuestas_grupo (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
                VALUES (%s, %s, %s, NULL, 0, 0, %s)
                ON DUPLICATE KEY UPDATE opcion_id = NULL, es_correcta = 0, puntos = 0, tiempo_respuesta = %s
            """, (sesion_id, g.user['id'], question_id, tiempo_respuesta, tiempo_respuesta))

            db.commit()

            # Actualizar pregunta actual del usuario
            cursor.execute("""
                UPDATE usuario_estado_grupo
                SET pregunta_actual = %s
                WHERE sesion_id = %s AND user_id = %s
            """, (question_id, sesion_id, g.user['id']))

            db.commit()

            return jsonify({
                'success': True,
                'correct': False,
                'points': 0,
                'timeout': True
            })

        # Obtener información de la opción y la pregunta
        cursor.execute("""
            SELECT o.es_correcta, p.puntos, p.tiempo_limite
            FROM opciones_respuesta o
            JOIN preguntas p ON o.pregunta_id = p.id
            WHERE o.id = %s AND o.pregunta_id = %s
        """, (answer_id, question_id))

        resultado = cursor.fetchone()
        if not resultado:
            return jsonify({'success': False, 'message': 'Opción no encontrada'})

        # Calcular puntos (más rápido = más puntos)
        es_correcta = resultado['es_correcta']
        puntos = 0
        if es_correcta:
            puntos_base = resultado['puntos'] or 1000
            tiempo_limite = resultado['tiempo_limite'] or 30
            # Fórmula: puntaje = puntaje_base * (tiempo_restante / tiempo_total)
            tiempo_restante = max(0, tiempo_limite - tiempo_respuesta)
            factor_tiempo = tiempo_restante / tiempo_limite
            # Redondear hacia arriba para que siempre se otorgue al menos 1 punto
            puntos = max(1, math.ceil(puntos_base * factor_tiempo))

        # Guardar respuesta
        cursor.execute("""
            INSERT INTO respuestas_grupo (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE opcion_id = %s, es_correcta = %s, puntos = %s, tiempo_respuesta = %s
        """, (sesion_id, g.user['id'], question_id, answer_id, es_correcta, puntos, tiempo_respuesta,
              answer_id, es_correcta, puntos, tiempo_respuesta))

        db.commit()

        # Actualizar pregunta actual del usuario
        cursor.execute("""
            UPDATE usuario_estado_grupo
            SET pregunta_actual = %s
            WHERE sesion_id = %s AND user_id = %s
        """, (question_id, sesion_id, g.user['id']))

        db.commit()

        return jsonify({
            'success': True,
            'correct': bool(es_correcta),
            'points': puntos
        })

    except Exception as e:
        if db:
            db.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/api/grupo/pregunta-estado/<int:sesion_id>/<int:pregunta_id>')
@jwt_required()
def api_grupo_pregunta_estado(sesion_id, pregunta_id):
    """API para verificar si todos respondieron una pregunta"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Contar total de usuarios en la sesión (EXCLUYENDO al creador/host)
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM usuario_estado_grupo ues
            JOIN sesiones_grupo sg ON ues.sesion_id = sg.id
            WHERE ues.sesion_id = %s AND ues.user_id != sg.iniciado_por
        """, (sesion_id,))
        total_usuarios = cursor.fetchone()['total']

        # Contar cuántos ya respondieron esta pregunta
        cursor.execute("""
            SELECT COUNT(*) as respondidos
            FROM respuestas_grupo
            WHERE sesion_id = %s AND pregunta_id = %s
        """, (sesion_id, pregunta_id))
        respondidos = cursor.fetchone()['respondidos']

        todos_respondieron = (total_usuarios > 0 and respondidos >= total_usuarios)

        return jsonify({
            'success': True,
            'all_answered': todos_respondieron,
            'total': total_usuarios,
            'answered': respondidos
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@app.route('/api/grupo/obtener-pregunta-actual/<int:sesion_id>')
@jwt_required()
def api_grupo_obtener_pregunta_actual(sesion_id):
    """API para obtener el índice de la pregunta actual del usuario o del host"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Primero verificar la información de la sesión
        cursor.execute("""
            SELECT sg.cuestionario_id, sg.iniciado_por, sg.pregunta_actual_id
            FROM sesiones_grupo sg
            WHERE sg.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()
        if not sesion:
            return jsonify({'success': False, 'message': 'Sesión no encontrada'})

        # Obtener todas las preguntas ordenadas
        cursor.execute("""
            SELECT id
            FROM preguntas
            WHERE cuestionario_id = %s
            ORDER BY orden ASC
        """, (sesion['cuestionario_id'],))

        preguntas = cursor.fetchall()

        # Si es el creador (host), usar pregunta_actual_id de sesiones_grupo
        if sesion['iniciado_por'] == g.user['id']:
            if sesion['pregunta_actual_id'] is not None:
                # Buscar el índice de la pregunta actual
                for idx, pregunta in enumerate(preguntas):
                    if pregunta['id'] == sesion['pregunta_actual_id']:
                        return jsonify({
                            'success': True,
                            'pregunta_id': sesion['pregunta_actual_id'],
                            'pregunta_index': idx,
                            'total_preguntas': len(preguntas)
                        })
            # Si no hay pregunta actual, empezar desde 0
            return jsonify({
                'success': True,
                'pregunta_id': None,
                'pregunta_index': 0,
                'total_preguntas': len(preguntas)
            })

        # Para estudiantes: Verificar si el usuario está en la sesión
        cursor.execute("""
            SELECT pregunta_actual
            FROM usuario_estado_grupo
            WHERE sesion_id = %s AND user_id = %s
        """, (sesion_id, g.user['id']))

        resultado = cursor.fetchone()

        if not resultado:
            # Usuario no registrado, empezar desde 0
            return jsonify({
                'success': True,
                'pregunta_id': None,
                'pregunta_index': 0,
                'total_preguntas': len(preguntas)
            })

        # CORRECCIÓN: Usar pregunta_actual_id de sesiones_grupo para sincronizar todos
        # Esto asegura que todos los usuarios estén en la misma pregunta que el host
        if sesion['pregunta_actual_id'] is not None:
            # Buscar el índice de la pregunta actual de la sesión
            for idx, pregunta in enumerate(preguntas):
                if pregunta['id'] == sesion['pregunta_actual_id']:
                    return jsonify({
                        'success': True,
                        'pregunta_id': sesion['pregunta_actual_id'],
                        'pregunta_index': idx,
                        'total_preguntas': len(preguntas)
                    })

        # Si no hay pregunta actual en la sesión, buscar la última pregunta del usuario
        if resultado['pregunta_actual'] is not None:
            pregunta_id_actual = resultado['pregunta_actual']
            # Buscar el índice de esta pregunta
            for idx, pregunta in enumerate(preguntas):
                if pregunta['id'] == pregunta_id_actual:
                    # CORRECCIÓN: NO avanzar automáticamente a la siguiente
                    # El usuario debe quedarse en la pregunta actual hasta que termine el tiempo
                    return jsonify({
                        'success': True,
                        'pregunta_id': pregunta_id_actual,
                        'pregunta_index': idx,
                        'total_preguntas': len(preguntas)
                    })

        # Buscar la última pregunta que respondió
        cursor.execute("""
            SELECT pregunta_id
            FROM respuestas_grupo
            WHERE sesion_id = %s AND user_id = %s
            ORDER BY created_at DESC
            LIMIT 1
        """, (sesion_id, g.user['id']))

        ultima_respuesta = cursor.fetchone()

        if ultima_respuesta:
            pregunta_id_actual = ultima_respuesta['pregunta_id']
            # Buscar el índice de esta pregunta
            for idx, pregunta in enumerate(preguntas):
                if pregunta['id'] == pregunta_id_actual:
                    # CORRECCIÓN: Quedarse en la pregunta que respondió, no avanzar
                    return jsonify({
                        'success': True,
                        'pregunta_id': pregunta_id_actual,
                        'pregunta_index': idx,
                        'total_preguntas': len(preguntas)
                    })

        # No ha respondido nada, empezar desde 0
        return jsonify({
            'success': True,
            'pregunta_id': None,
            'pregunta_index': 0,
            'total_preguntas': len(preguntas)
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/api/grupo/timer-sync/<int:sesion_id>', methods=['GET'])
@jwt_required()
def api_grupo_timer_sync(sesion_id):
    """API para sincronizar el timer entre todos los jugadores"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información del timer actual de la sesión
        cursor.execute("""
            SELECT
                pregunta_actual_id,
                pregunta_inicio_time,
                pregunta_tiempo_limite
            FROM sesiones_grupo
            WHERE id = %s
        """, (sesion_id,))

        sesion_info = cursor.fetchone()

        if not sesion_info or not sesion_info['pregunta_inicio_time']:
            return jsonify({
                'success': False,
                'message': 'No hay pregunta activa',
                'tiempo_restante': 0
            })

        # Calcular tiempo transcurrido
        from datetime import datetime
        inicio = sesion_info['pregunta_inicio_time']
        ahora = datetime.now()
        tiempo_transcurrido = (ahora - inicio).total_seconds()
        tiempo_limite = sesion_info['pregunta_tiempo_limite'] or 30
        tiempo_restante = max(0, int(tiempo_limite - tiempo_transcurrido))

        return jsonify({
            'success': True,
            'tiempo_restante': tiempo_restante,
            'pregunta_id': sesion_info['pregunta_actual_id'],
            'tiempo_limite': tiempo_limite
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/api/grupo/set-pregunta-actual', methods=['POST'])
@jwt_required()
def api_grupo_set_pregunta_actual():
    """API para que el docente establezca la pregunta actual (inicia el timer compartido)"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')
    pregunta_id = data.get('pregunta_id')
    tiempo_limite = data.get('tiempo_limite', 30)

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar que el usuario es el creador de la sesión
        cursor.execute("""
            SELECT iniciado_por
            FROM sesiones_grupo
            WHERE id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()
        if not sesion or sesion['iniciado_por'] != g.user['id']:
            return jsonify({'success': False, 'message': 'No autorizado - solo el creador puede controlar el timer'}), 403

        # Actualizar la pregunta actual y el timestamp de inicio
        from datetime import datetime
        cursor.execute("""
            UPDATE sesiones_grupo
            SET pregunta_actual_id = %s,
                pregunta_inicio_time = %s,
                pregunta_tiempo_limite = %s
            WHERE id = %s
        """, (pregunta_id, datetime.now(), tiempo_limite, sesion_id))

        db.commit()

        return jsonify({
            'success': True,
            'message': 'Timer sincronizado iniciado'
        })

    except Exception as e:
        if db:
            db.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/api/grupo/start-game', methods=['POST'])
@jwt_required()
def api_start_game():
    """API para que el creador inicie el juego (sin necesidad de que estén listos)"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar que el usuario es el creador de la sesión
        cursor.execute("""
            SELECT iniciado_por FROM sesiones_grupo WHERE id = %s
        """, (sesion_id,))
        sesion_info = cursor.fetchone()

        if not sesion_info or g.user['id'] != sesion_info['iniciado_por']:
            cursor.close()
            db.close()
            return jsonify({'success': False, 'message': 'Solo el creador puede iniciar el juego'})

        # Contar participantes (sin importar si están listos o no)
        cursor.execute("""
            SELECT COUNT(*) as total_participantes
            FROM usuario_estado_grupo ues
            JOIN sesiones_grupo sg ON ues.sesion_id = sg.id
            WHERE ues.sesion_id = %s AND ues.user_id != sg.iniciado_por
        """, (sesion_id,))
        participantes = cursor.fetchone()['total_participantes']

        # El creador puede iniciar aunque no haya participantes
        # (útil para probar el juego solo)

        # Iniciar el juego
        cursor.execute("""
            UPDATE sesiones_grupo
            SET estado = 'en_progreso', started_at = NOW()
            WHERE id = %s
        """, (sesion_id,))

        db.commit()
        cursor.close()
        db.close()

        mensaje = f'Juego iniciado con {participantes} participante(s)' if participantes > 0 else 'Juego iniciado (modo prueba sin participantes)'

        return jsonify({
            'success': True,
            'message': mensaje,
            'participantes': participantes
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/grupo/finalizar-sesion', methods=['POST'])
@jwt_required()
def api_finalizar_sesion():
    """API para marcar la sesión como finalizada y asignar recompensas"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Actualizar estado de la sesión
        cursor.execute("""
            UPDATE sesiones_grupo
            SET estado = 'finalizado', finished_at = NOW()
            WHERE id = %s
        """, (sesion_id,))

        db.commit()

        # Asignar puntosmoneda a los ganadores
        asignar_puntosmoneda_grupo(sesion_id)

        return jsonify({'success': True})

    except Exception as e:
        if db:
            db.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/grupo/resultados/<int:sesion_id>')
def grupo_resultados(sesion_id):
    """Página de resultados finales de la sesión de grupo"""
    if not g.user:
        flash('Debes iniciar sesión para ver los resultados', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión incluyendo el iniciador
        cursor.execute("""
            SELECT sg.id, sg.grupo_id, sg.cuestionario_id, sg.iniciado_por,
                   g.nombre as grupo_nombre, c.titulo as cuestionario_titulo
            FROM sesiones_grupo sg
            JOIN grupos g ON sg.grupo_id = g.id
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            WHERE sg.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('grupos'))

        # Verificar si el usuario actual es el creador/anfitrión
        es_creador = (g.user['id'] == sesion['iniciado_por'])

        # Obtener el total de preguntas
        cursor.execute("""
            SELECT COUNT(p.id) as total_preguntas
            FROM preguntas p
            WHERE p.cuestionario_id = %s
        """, (sesion['cuestionario_id'],))
        total_preguntas = cursor.fetchone()['total_preguntas']

        # Obtener resultados de todos los participantes
        cursor.execute("""
            SELECT
                u.id,
                u.username,
                COALESCE(SUM(rg.puntos), 0) as score,
                COALESCE(SUM(rg.es_correcta), 0) as correct_answers,
                COUNT(DISTINCT rg.pregunta_id) as preguntas_respondidas
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            LEFT JOIN respuestas_grupo rg ON rg.user_id = u.id AND rg.sesion_id = %s
            WHERE ues.sesion_id = %s
            GROUP BY u.id, u.username
            ORDER BY score DESC, correct_answers DESC
        """, (sesion_id, sesion_id))
        resultados_raw = cursor.fetchall()

        # Limpiar datos de resultados para asegurar serialización JSON
        resultados = []
        for r in resultados_raw:
            resultados.append({
                'id': int(r.get('id', 0)) if r.get('id') is not None else 0,
                'username': str(r.get('username', 'Usuario desconocido')),
                'score': int(r.get('score', 0)) if r.get('score') is not None else 0,
                'correct_answers': int(r.get('correct_answers', 0)) if r.get('correct_answers') is not None else 0,
                'preguntas_respondidas': int(r.get('preguntas_respondidas', 0)) if r.get('preguntas_respondidas') is not None else 0
            })

        # Obtener detalle de respuestas para el template (solo si es necesario)
        detalle_respuestas = []
        if es_creador or g.user['id'] in [r['id'] for r in resultados]:
            # Obtener todas las respuestas de esta sesión
            cursor.execute("""
                SELECT
                    rg.user_id,
                    rg.pregunta_id,
                    p.texto_pregunta,
                    p.orden,
                    p.puntos as puntos_pregunta,
                    rg.es_correcta,
                    rg.puntos as puntos_obtenidos,
                    rg.tiempo_respuesta,
                    op_seleccionada.texto_opcion as texto_respuesta,
                    op_correcta.texto_opcion as texto_correcta,
                    CASE WHEN rg.opcion_id IS NOT NULL THEN 1 ELSE 0 END as opcion_seleccionada
                FROM respuestas_grupo rg
                JOIN preguntas p ON rg.pregunta_id = p.id
                LEFT JOIN opciones_respuesta op_seleccionada ON rg.opcion_id = op_seleccionada.id
                LEFT JOIN opciones_respuesta op_correcta ON op_correcta.pregunta_id = p.id AND op_correcta.es_correcta = 1
                WHERE rg.sesion_id = %s
                ORDER BY rg.user_id, p.orden
            """, (sesion_id,))
            detalle_raw = cursor.fetchall()

            # Limpiar datos de detalle
            for d in detalle_raw:
                detalle_respuestas.append({
                    'user_id': int(d.get('user_id', 0)) if d.get('user_id') is not None else 0,
                    'pregunta_id': int(d.get('pregunta_id', 0)) if d.get('pregunta_id') is not None else 0,
                    'texto_pregunta': str(d.get('texto_pregunta', '')) if d.get('texto_pregunta') is not None else '',
                    'orden': int(d.get('orden', 0)) if d.get('orden') is not None else 0,
                    'puntos_pregunta': int(d.get('puntos_pregunta', 0)) if d.get('puntos_pregunta') is not None else 0,
                    'es_correcta': bool(d.get('es_correcta')) if d.get('es_correcta') is not None else False,
                    'puntos_obtenidos': int(d.get('puntos_obtenidos', 0)) if d.get('puntos_obtenidos') is not None else 0,
                    'tiempo_respuesta': float(d.get('tiempo_respuesta', 0)) if d.get('tiempo_respuesta') is not None else 0.0,
                    'texto_respuesta': str(d.get('texto_respuesta', '')) if d.get('texto_respuesta') is not None else '',
                    'texto_correcta': str(d.get('texto_correcta', '')) if d.get('texto_correcta') is not None else '',
                    'opcion_seleccionada': bool(d.get('opcion_seleccionada')) if d.get('opcion_seleccionada') is not None else False
                })

        cursor.close()
        db.close()

        # Logging para depuración
        import sys
        sys.stderr.write(f"\n=== DEBUG grupo_resultados ===\n")
        sys.stderr.write(f"Sesión ID: {sesion_id}\n")
        sys.stderr.write(f"Resultados count: {len(resultados)}\n")
        sys.stderr.write(f"Detalle respuestas count: {len(detalle_respuestas)}\n")
        sys.stderr.write(f"Total preguntas: {total_preguntas}\n")
        sys.stderr.flush()

        # Intentar renderizar el template con manejo de errores
        try:
            # Asegurar que detalle_respuestas sea serializable
            detalle_respuestas_clean = json.dumps(detalle_respuestas, default=str, ensure_ascii=False)
            sys.stderr.write(f"Detalle respuestas serializado OK\n")
            sys.stderr.flush()
        except Exception as json_err:
            sys.stderr.write(f"ERROR al serializar detalle_respuestas: {str(json_err)}\n")
            sys.stderr.write(f"Tipo de detalle_respuestas: {type(detalle_respuestas)}\n")
            if detalle_respuestas:
                sys.stderr.write(f"Primer elemento tipo: {type(detalle_respuestas[0]) if detalle_respuestas else 'N/A'}\n")
                if detalle_respuestas:
                    sys.stderr.write(f"Primer elemento keys: {list(detalle_respuestas[0].keys()) if isinstance(detalle_respuestas[0], dict) else 'N/A'}\n")
            sys.stderr.flush()
            detalle_respuestas = []  # Usar lista vacía si falla

        try:
            template_result = render_template('grupo_resultados.html',
                                 sesion_id=sesion_id,
                                 grupo_nombre=sesion.get('grupo_nombre', 'Grupo desconocido'),
                                 cuestionario_titulo=sesion.get('cuestionario_titulo', 'Cuestionario'),
                                 resultados=resultados,
                                 total_preguntas=int(total_preguntas) if total_preguntas is not None else 0,
                                 es_creador=es_creador,
                                 detalle_respuestas=detalle_respuestas)
            sys.stderr.write(f"Template renderizado OK\n")
            sys.stderr.flush()
            return template_result
        except Exception as render_err:
            import traceback
            error_traceback = traceback.format_exc()
            sys.stderr.write(f"\n=== ERROR AL RENDERIZAR TEMPLATE ===\n")
            sys.stderr.write(f"Error: {str(render_err)}\n")
            sys.stderr.write(f"Traceback:\n{error_traceback}\n")
            sys.stderr.flush()
            raise  # Re-lanzar para que sea capturado por el except externo

    except Exception as e:
        import sys
        import traceback

        if db:
            try:
                cursor.close()
            except:
                pass
            try:
                db.close()
            except:
                pass

        error_msg = f'Error al cargar resultados: {str(e)}'
        error_traceback = traceback.format_exc()

        # SIEMPRE escribir en stderr (esto siempre funciona)
        sys.stderr.write(f"\n{'='*70}\n")
        sys.stderr.write(f"ERROR EN grupo_resultados - {datetime.now().isoformat()}\n")
        sys.stderr.write(f"{'='*70}\n")
        sys.stderr.write(f"Error: {error_msg}\n")
        sys.stderr.write(f"Traceback completo:\n{error_traceback}\n")
        sys.stderr.write(f"{'='*70}\n")
        sys.stderr.flush()

        # Intentar escribir en archivo de log (opcional, no crítico)
        try:
            import os
            log_path = os.path.join(os.path.dirname(__file__), 'error_log.txt')
            with open(log_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"\n{'='*70}\n")
                log_file.write(f"ERROR EN grupo_resultados - {datetime.now().isoformat()}\n")
                log_file.write(f"{'='*70}\n")
                log_file.write(f"Error: {error_msg}\n")
                log_file.write(f"Traceback completo:\n{error_traceback}\n")
                log_file.write(f"{'='*70}\n")
        except Exception as log_error:
            # Si falla escribir en archivo, no importa, ya está en stderr
            sys.stderr.write(f"Nota: No se pudo escribir en archivo de log: {log_error}\n")
            sys.stderr.flush()

        flash(error_msg, 'error')
        return redirect(url_for('grupos'))

@app.route('/api/grupo/results/<int:sesion_id>')
@jwt_required()
def api_grupo_results(sesion_id):
    """API para obtener resultados del grupo"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener el total de preguntas del cuestionario de esta sesión
        cursor.execute("""
            SELECT COUNT(p.id) as total_preguntas
            FROM sesiones_grupo sg
            JOIN preguntas p ON sg.cuestionario_id = p.cuestionario_id
            WHERE sg.id = %s
        """, (sesion_id,))
        total_preguntas_row = cursor.fetchone()
        total_preguntas = total_preguntas_row['total_preguntas'] if total_preguntas_row else 0

        # Obtener puntuaciones y aciertos de todos los usuarios
        cursor.execute("""
            SELECT
                u.username,
                SUM(rg.puntos) as score,
                SUM(rg.es_correcta) as correct_answers
            FROM respuestas_grupo rg
            JOIN users u ON rg.user_id = u.id
            WHERE rg.sesion_id = %s
            GROUP BY u.id, u.username
            ORDER BY score DESC
        """, (sesion_id,))
        resultados = cursor.fetchall()

        # Añadir el total de preguntas a cada resultado
        for r in resultados:
            r['total_questions'] = total_preguntas

        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'results': resultados
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/grupo/resultados/<int:sesion_id>/compartir-creador', methods=['POST'])
@jwt_required()
def api_grupo_resultados_compartir_creador(sesion_id):
    """Envía un correo al creador de la sesión con el link de la carpeta de Drive donde se guardan los resultados."""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener datos del creador y contexto del quiz/grupo
        cursor.execute(
            """
            SELECT sg.iniciado_por, u.email, u.username,
                   c.titulo AS cuestionario_titulo, COALESCE(g.nombre, 'Sala Temporal') AS grupo_nombre
            FROM sesiones_grupo sg
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            LEFT JOIN grupos g ON sg.grupo_id = g.id
            JOIN users u ON u.id = sg.iniciado_por
            WHERE sg.id = %s
            """,
            (sesion_id,)
        )
        row = cursor.fetchone()

        if not row:
            cursor.close(); db.close()
            return jsonify({'success': False, 'message': 'Sesión no encontrada'}), 404

        creator_email = row['email']
        creator_username = row['username']
        quiz_title = row['cuestionario_titulo']
        group_name = row['grupo_nombre']

        if not creator_email:
            cursor.close(); db.close()
            return jsonify({'success': False, 'message': 'El creador no tiene un correo configurado'}), 400

        # Link a la carpeta de Google Drive donde se suben los resultados
        folder_id = GOOGLE_DRIVE_FOLDER_ID
        folder_url = f"https://drive.google.com/drive/folders/{folder_id}"

        # Construir y enviar correo
        try:
            subject = f"Resultados disponibles - {quiz_title}"
            body = (
                f"Hola {creator_username},\n\n"
                f"Los resultados del cuestionario '{quiz_title}' (grupo: '{group_name}') ya están disponibles.\n\n"
                f"Puedes acceder a la carpeta donde se guardan los archivos de resultados en Google Drive aquí:\n"
                f"{folder_url}\n\n"
                f"Sugerencia: Verifica que tu cuenta tenga acceso a esa carpeta.\n\n"
                f"Saludos,\nEquipo RoBot"
            )

            msg = Message(
                subject,
                sender=app.config['MAIL_USERNAME'],
                recipients=[creator_email]
            )
            msg.body = body
            mail.send(msg)
        except Exception as e:
            # Fallback: devolver éxito con enlace para copiar manualmente
            cursor.close(); db.close()
            return jsonify({
                'success': True,
                'emailed': False,
                'folder_url': folder_url,
                'message': f'No se pudo enviar el correo automáticamente. Comparte este enlace: {folder_url}'
            })

        cursor.close(); db.close()
        return jsonify({'success': True, 'emailed': True, 'message': 'Correo enviado al creador con el enlace de la carpeta'})

    except Exception as e:
        try:
            db.close()
        except:
            pass
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/resultados/<pin>')
def resultados_quiz(pin):
    """Página de resultados para una sesión de juego individual."""
    if not g.user:
        flash('Debes iniciar sesión para ver los resultados.', 'warning')
        return redirect(url_for('login'))

    db = None
    game_data = {
        "totalTime": 0,
        "questions": [],
        "participants": []
    }

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # 1. Obtener el cuestionario por PIN
        cursor.execute("""
            SELECT id, titulo, created_at FROM cuestionarios WHERE pin = %s
        """, (pin,))
        quiz_info = cursor.fetchone()

        if not quiz_info:
            flash('Cuestionario con ese PIN no encontrado.', 'error')
            return redirect(url_for('home'))

        cuestionario_id = quiz_info['id']

        # 2. Obtener todas las preguntas del cuestionario para el JSON
        cursor.execute("""
            SELECT id, texto_pregunta as text, orden
            FROM preguntas
            WHERE cuestionario_id = %s ORDER BY orden
        """, (cuestionario_id,))
        preguntas_raw = cursor.fetchall()
        # Limpiar datos de preguntas para asegurar serialización JSON
        game_data['questions'] = []
        for p in preguntas_raw:
            pregunta_id = p.get('id')
            texto = p.get('text')
            orden = p.get('orden')

            game_data['questions'].append({
                "id": int(pregunta_id) if pregunta_id is not None else 0,
                "text": str(texto) if texto is not None else "",
                "orden": int(orden) if orden is not None else 0
            })

        # 3. Obtener todas las sesiones de juego para este cuestionario
        cursor.execute("""
            SELECT id, user_id, fecha_inicio, fecha_fin FROM sesiones_juego
            WHERE cuestionario_id = %s AND user_id IS NOT NULL
        """, (cuestionario_id,))
        sesiones = cursor.fetchall()

        if not sesiones:
            flash('Aún no se ha jugado ninguna partida para este cuestionario.', 'info')
            return redirect(url_for('quiz_details', cuestionario_id=cuestionario_id))

        # 4. Procesar los resultados de todas las sesiones
        participants_map = {}
        total_game_duration = 0

        for sesion in sesiones:
            # Calcular duración total del juego
            if sesion.get('fecha_inicio') and sesion.get('fecha_fin'):
                duration = (sesion['fecha_fin'] - sesion['fecha_inicio']).total_seconds()
                total_game_duration = max(total_game_duration, duration)

            # Obtener respuestas de esta sesión
            sesion_id = sesion.get('id')
            if not sesion_id:
                continue

            cursor.execute("""
                SELECT
                    rp.user_id,
                    u.username,
                    rp.pregunta_id,
                    rp.es_correcta,
                    rp.puntos_obtenidos,
                    rp.tiempo_respuesta,
                    op.texto_opcion as choice
                FROM respuestas_participantes rp
                JOIN users u ON rp.user_id = u.id
                LEFT JOIN opciones_respuesta op ON rp.opcion_id = op.id
                WHERE rp.sesion_juego_id = %s
            """, (sesion_id,))
            respuestas = cursor.fetchall()

            for r in respuestas:
                user_id = r.get('user_id')
                if user_id is None:
                    continue

                if user_id not in participants_map:
                    participants_map[user_id] = {
                        "id": int(user_id) if user_id is not None else 0,
                        "name": str(r.get('username', 'Usuario desconocido')),
                        "answers": []
                    }

                # Limpiar cada valor antes de agregarlo
                pregunta_id = r.get('pregunta_id')
                es_correcta = r.get('es_correcta')
                puntos = r.get('puntos_obtenidos')
                tiempo = r.get('tiempo_respuesta')
                choice = r.get('choice')

                participants_map[user_id]['answers'].append({
                    "questionId": int(pregunta_id) if pregunta_id is not None else 0,
                    "correct": bool(es_correcta) if es_correcta is not None else False,
                    "points": int(puntos) if puntos is not None else 0,
                    "time": float(tiempo) if tiempo is not None else 0.0,
                    "choice": str(choice) if choice is not None else "N/A"
                })

        game_data['participants'] = list(participants_map.values())
        game_data['totalTime'] = float(total_game_duration) if total_game_duration is not None else 0.0

        # Función helper para limpiar datos antes de serializar
        def clean_for_json(obj):
            import decimal
            from datetime import datetime, date, timedelta

            if obj is None:
                return None
            elif isinstance(obj, bool):
                return obj
            elif isinstance(obj, (int, float)):
                return obj
            elif isinstance(obj, str):
                return obj
            elif isinstance(obj, decimal.Decimal):
                return float(obj)
            elif isinstance(obj, (datetime, date)):
                return obj.isoformat()
            elif isinstance(obj, timedelta):
                return obj.total_seconds()
            elif isinstance(obj, dict):
                return {str(k): clean_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [clean_for_json(item) for item in obj]
            elif hasattr(obj, '__dict__'):
                return clean_for_json(obj.__dict__)
            else:
                try:
                    return str(obj)
                except:
                    return None

        game_data_clean = clean_for_json(game_data)

        cursor.close()
        db.close()

        # Intentar serializar y capturar errores específicos
        try:
            game_data_json = json.dumps(game_data_clean, default=str, ensure_ascii=False)
        except TypeError as e:
            # Si hay un error de serialización, intentar con una función default más robusta
            def json_default(obj):
                import decimal
                from datetime import datetime, date, timedelta
                if isinstance(obj, decimal.Decimal):
                    return float(obj)
                elif isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                elif isinstance(obj, timedelta):
                    return obj.total_seconds()
                elif hasattr(obj, '__dict__'):
                    return obj.__dict__
                else:
                    return str(obj)

            try:
                game_data_json = json.dumps(game_data_clean, default=json_default, ensure_ascii=False)
            except Exception as e2:
                # Si aún falla, registrar el error y usar un objeto vacío
                import traceback
                error_msg = f"Error al serializar JSON: {str(e2)}\n"
                error_msg += f"Traceback: {traceback.format_exc()}\n"
                error_msg += f"game_data_clean type: {type(game_data_clean)}\n"
                error_msg += f"game_data_clean keys: {game_data_clean.keys() if isinstance(game_data_clean, dict) else 'N/A'}\n"

                # Escribir en archivo de log y también imprimir
                try:
                    with open('error_log.txt', 'a', encoding='utf-8') as f:
                        f.write(f"\n{'='*50}\n")
                        f.write(f"Error en resultados_quiz - {datetime.now().isoformat()}\n")
                        f.write(f"{'='*50}\n")
                        f.write(error_msg)
                except Exception as log_error:
                    print(f"No se pudo escribir en log: {log_error}")

                print(error_msg)
                game_data_json = json.dumps({"error": "Error al cargar datos", "totalTime": 0, "questions": [], "participants": []})

        return render_template('resultados_quiz.html',
                               title=f"Resultados de {quiz_info['titulo']}",
                               game_data_json=game_data_json)

    except Exception as e:
        if db:
            try:
                cursor.close()
            except:
                pass
            db.close()
        import traceback
        error_msg = f'Error al cargar los resultados: {str(e)}'
        error_traceback = traceback.format_exc()

        # Escribir en archivo de log y también imprimir
        try:
            with open('error_log.txt', 'a', encoding='utf-8') as f:
                f.write(f"\n{'='*50}\n")
                f.write(f"Error en resultados_quiz - {datetime.now().isoformat()}\n")
                f.write(f"{'='*50}\n")
                f.write(f"Error: {error_msg}\n")
                f.write(f"Traceback completo:\n{error_traceback}\n")
        except Exception as log_error:
            print(f"No se pudo escribir en log: {log_error}")

        print(f"Error completo: {error_traceback}")
        flash(error_msg, 'error')
        return redirect(url_for('home'))

# =====================
#  Sesiones Individuales (Sistema similar a grupos)
# =====================

@app.route('/api/individual/iniciar-cuestionario', methods=['POST'])
@jwt_required()
def api_iniciar_cuestionario_individual():
    """API para iniciar un cuestionario individual con sistema de sala"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    # Solo docentes pueden iniciar cuestionarios individuales
    if g.user.get('role') != 'docente':
        return jsonify({'success': False, 'message': 'Solo los docentes pueden iniciar cuestionarios'}), 403

    data = request.get_json()
    cuestionario_id = data.get('cuestionario_id')

    if not cuestionario_id:
        return jsonify({'success': False, 'message': 'ID de cuestionario requerido'}), 400

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar que el cuestionario existe
        cursor.execute("SELECT id, titulo FROM cuestionarios WHERE id = %s", (cuestionario_id,))
        cuestionario = cursor.fetchone()

        if not cuestionario:
            return jsonify({'success': False, 'message': 'Cuestionario no encontrado'})

        # Generar un session_code único de 6 caracteres
        import string
        import random
        while True:
            session_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            # Verificar que no exista en sesiones_grupo ni sesiones_individual
            cursor.execute("SELECT id FROM sesiones_grupo WHERE session_code = %s", (session_code,))
            if cursor.fetchone():
                continue
            cursor.execute("SELECT id FROM sesiones_individual WHERE session_code = %s", (session_code,))
            if not cursor.fetchone():
                break

        # Crear sesión individual
        cursor.execute("""
            INSERT INTO sesiones_individual (cuestionario_id, iniciado_por, session_code, created_at)
            VALUES (%s, %s, %s, NOW())
        """, (cuestionario_id, g.user['id'], session_code))

        sesion_id = cursor.lastrowid
        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'sesion_id': sesion_id,
            'session_code': session_code,
            'message': 'Sesión creada exitosamente'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al iniciar cuestionario: {str(e)}'})

@app.route('/individual/quiz/<int:sesion_id>')
def individual_quiz(sesion_id):
    """Página de sala de espera para cuestionario individual"""
    if not g.user:
        flash('Debes iniciar sesión para participar', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión
        cursor.execute("""
            SELECT si.id, si.cuestionario_id, si.estado, si.session_code,
                   c.titulo, c.pin, c.descripcion
            FROM sesiones_individual si
            JOIN cuestionarios c ON si.cuestionario_id = c.id
            WHERE si.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('my_quizzes'))

        # Si la sesión ya finalizó, redirigir
        if sesion['estado'] == 'finalizado':
            flash('Esta sesión ya ha finalizado', 'info')
            return redirect(url_for('my_quizzes'))

        # Registrar automáticamente al usuario en la sesión si no está ya
        cursor.execute("""
            INSERT INTO usuario_estado_individual (sesion_id, user_id, esta_listo)
            VALUES (%s, %s, 0)
            ON DUPLICATE KEY UPDATE user_id = user_id
        """, (sesion_id, g.user['id']))

        db.commit()

        # Obtener todos los usuarios en la sala
        cursor.execute("""
            SELECT u.id, u.username, uei.esta_listo
            FROM usuario_estado_individual uei
            JOIN users u ON uei.user_id = u.id
            WHERE uei.sesion_id = %s
            ORDER BY uei.created_at ASC
        """, (sesion_id,))

        participantes = cursor.fetchall()

        # Obtener el número de preguntas
        cursor.execute("""
            SELECT COUNT(*) as count FROM preguntas WHERE cuestionario_id = %s
        """, (sesion['cuestionario_id'],))

        preguntas_count = cursor.fetchone()['count']

        cursor.close()
        db.close()

        cuestionario_info = {
            'titulo': sesion['titulo'],
            'pin': sesion['pin'],
            'descripcion': sesion['descripcion'],
            'preguntas_count': preguntas_count
        }

        return render_template('individual_quiz.html',
                             sesion_id=sesion_id,
                             session_code=sesion['session_code'],
                             cuestionario=cuestionario_info,
                             participantes=participantes,
                             user_id=g.user['id'])

    except Exception as e:
        flash(f'Error al cargar la sala: {str(e)}', 'error')
        return redirect(url_for('my_quizzes'))

@app.route('/api/individual/unirse-sesion', methods=['POST'])
@jwt_required()
def api_unirse_sesion_individual():
    """API para unirse a una sesión individual usando el código de sala"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    session_code = data.get('session_code', '').strip().upper()

    if len(session_code) != 6:
        return jsonify({'success': False, 'message': 'El código debe tener 6 caracteres'})

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Buscar sesión por código
        cursor.execute("""
            SELECT id, estado FROM sesiones_individual
            WHERE session_code = %s
        """, (session_code,))

        sesion = cursor.fetchone()

        if not sesion:
            return jsonify({'success': False, 'message': 'Código de sala inválido'})

        if sesion['estado'] == 'finalizado':
            return jsonify({'success': False, 'message': 'Esta sesión ya ha finalizado'})

        # Registrar al usuario en la sesión
        cursor.execute("""
            INSERT INTO usuario_estado_individual (sesion_id, user_id, esta_listo)
            VALUES (%s, %s, 0)
            ON DUPLICATE KEY UPDATE user_id = user_id
        """, (sesion['id'], g.user['id']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'sesion_id': sesion['id'],
            'message': 'Te has unido a la sesión'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al unirse: {str(e)}'})

@app.route('/api/individual/ready', methods=['POST'])
@jwt_required()
def api_individual_ready():
    """API para marcar al usuario como listo en sesión individual"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Verificar que el usuario está en la sesión
        cursor.execute("""
            SELECT id FROM usuario_estado_individual
            WHERE sesion_id = %s AND user_id = %s
        """, (sesion_id, g.user['id']))

        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'No estás en esta sesión'})

        # Marcar como listo
        cursor.execute("""
            UPDATE usuario_estado_individual
            SET esta_listo = 1
            WHERE sesion_id = %s AND user_id = %s
        """, (sesion_id, g.user['id']))

        db.commit()

        # Verificar si todos están listos
        cursor.execute("""
            SELECT COUNT(*) AS total_en_sala
            FROM usuario_estado_individual
            WHERE sesion_id = %s
        """, (sesion_id,))

        total_en_sala = cursor.fetchone()['total_en_sala']

        cursor.execute("""
            SELECT COUNT(*) AS listos
            FROM usuario_estado_individual
            WHERE sesion_id = %s AND esta_listo = 1
        """, (sesion_id,))

        listos = cursor.fetchone()['listos']

        all_ready = (total_en_sala > 0 and listos == total_en_sala)

        if all_ready:
            # Actualizar estado de la sesión a 'en_progreso'
            cursor.execute("""
                UPDATE sesiones_individual
                SET estado = 'en_progreso', started_at = NOW()
                WHERE id = %s
            """, (sesion_id,))
            db.commit()

        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'all_ready': all_ready,
            'ready_count': listos,
            'total_count': total_en_sala,
            'message': 'Marcado como listo' if not all_ready else '¡Todos listos! Iniciando...'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/individual/juego/<int:sesion_id>')
def individual_juego(sesion_id):
    """Página para jugar el cuestionario individual (solo cuando está en progreso)"""
    if not g.user:
        flash('Debes iniciar sesión para participar', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión
        cursor.execute("""
            SELECT si.id, si.cuestionario_id, si.estado, si.session_code,
                   c.titulo, c.pin
            FROM sesiones_individual si
            JOIN cuestionarios c ON si.cuestionario_id = c.id
            WHERE si.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('my_quizzes'))

        # Verificar que el usuario está en la sesión
        cursor.execute("""
            SELECT id FROM usuario_estado_individual
            WHERE sesion_id = %s AND user_id = %s
        """, (sesion_id, g.user['id']))

        if not cursor.fetchone():
            flash('No estás registrado en esta sesión', 'error')
            return redirect(url_for('individual_quiz', sesion_id=sesion_id))

        # Verificar estado de la sesión
        if sesion['estado'] != 'en_progreso':
            flash('El juego aún no ha comenzado o ya finalizó', 'info')
            return redirect(url_for('individual_quiz', sesion_id=sesion_id))

        # Obtener preguntas del cuestionario
        cursor.execute("""
            SELECT id, texto_pregunta, tiempo_limite, puntos, orden
            FROM preguntas
            WHERE cuestionario_id = %s
            ORDER BY orden ASC
        """, (sesion['cuestionario_id'],))

        preguntas = cursor.fetchall()

        # Obtener opciones para cada pregunta
        preguntas_con_opciones = []
        for pregunta in preguntas:
            cursor.execute("""
                SELECT id, texto_opcion, es_correcta
                FROM opciones_respuesta
                WHERE pregunta_id = %s
                ORDER BY orden ASC
            """, (pregunta['id'],))

            opciones = cursor.fetchall()

            pregunta_dict = dict(pregunta)
            pregunta_dict['opciones'] = opciones
            preguntas_con_opciones.append(pregunta_dict)

        # Obtener participantes en la sala
        cursor.execute("""
            SELECT u.id, u.username
            FROM usuario_estado_individual uei
            JOIN users u ON uei.user_id = u.id
            WHERE uei.sesion_id = %s
        """, (sesion_id,))

        participantes = cursor.fetchall()

        cursor.close()
        db.close()

        return render_template('juego_individual.html',
                             sesion_id=sesion_id,
                             cuestionario={'titulo': sesion['titulo']},
                             preguntas_json=json.dumps(preguntas_con_opciones),
                             total_preguntas=len(preguntas),
                             participantes=participantes)

    except Exception as e:
        flash(f'Error al cargar el juego: {str(e)}', 'error')
        return redirect(url_for('my_quizzes'))

@app.route('/api/individual/participantes/<int:sesion_id>')
@jwt_required()
def api_individual_participantes(sesion_id):
    """API para obtener la lista actualizada de participantes en una sesión"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        cursor.execute("""
            SELECT u.id, u.username, uei.esta_listo as ready
            FROM usuario_estado_individual uei
            JOIN users u ON uei.user_id = u.id
            WHERE uei.sesion_id = %s
            ORDER BY uei.created_at ASC
        """, (sesion_id,))

        participantes = cursor.fetchall()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'participants': participantes
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/individual/status/<int:sesion_id>')
@jwt_required()
def api_individual_status(sesion_id):
    """API para obtener el estado de la sesión"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        cursor.execute("SELECT estado FROM sesiones_individual WHERE id = %s", (sesion_id,))
        sesion = cursor.fetchone()

        if not sesion:
            return jsonify({'success': False, 'message': 'Sesión no encontrada'})

        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'status': sesion['estado']
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/individual/answer', methods=['POST'])
@jwt_required()
def api_individual_answer():
    """API para enviar respuesta en sesión individual (igual que grupos)"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')
    question_id = data.get('question_id')
    answer_id = data.get('answer_id')  # Puede ser None para timeout
    tiempo_respuesta = data.get('tiempo_respuesta', 0)

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Si answer_id es None, es un timeout (sin respuesta)
        if answer_id is None or answer_id == 'null':
            # Timeout: registrar sin respuesta
            cursor.execute("""
                INSERT INTO respuestas_individual (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
                VALUES (%s, %s, %s, NULL, 0, 0, %s)
                ON DUPLICATE KEY UPDATE opcion_id = NULL, es_correcta = 0, puntos = 0, tiempo_respuesta = %s
            """, (sesion_id, g.user['id'], question_id, tiempo_respuesta, tiempo_respuesta))

            db.commit()

            # Actualizar pregunta actual del usuario
            cursor.execute("""
                UPDATE usuario_estado_individual
                SET pregunta_actual = %s
                WHERE sesion_id = %s AND user_id = %s
            """, (question_id, sesion_id, g.user['id']))

            db.commit()
            cursor.close()
            db.close()

            return jsonify({
                'success': True,
                'correct': False,
                'points': 0,
                'timeout': True
            })

        # Obtener información de la opción y la pregunta
        cursor.execute("""
            SELECT o.es_correcta, p.puntos, p.tiempo_limite
            FROM opciones_respuesta o
            JOIN preguntas p ON o.pregunta_id = p.id
            WHERE o.id = %s AND o.pregunta_id = %s
        """, (answer_id, question_id))

        resultado = cursor.fetchone()
        if not resultado:
            cursor.close()
            db.close()
            return jsonify({'success': False, 'message': 'Opción no encontrada'})

        # Calcular puntos (fórmula igual que grupos)
        es_correcta = resultado['es_correcta']
        puntos = 0
        if es_correcta:
            puntos_base = resultado['puntos'] or 1000
            tiempo_limite = resultado['tiempo_limite'] or 30
            # Más rápido = más puntos
            tiempo_restante = max(0, tiempo_limite - tiempo_respuesta)
            factor_tiempo = tiempo_restante / tiempo_limite
            # Redondear hacia arriba para que siempre se otorgue al menos 1 punto
            puntos = max(1, math.ceil(puntos_base * factor_tiempo))

        # Guardar respuesta en la tabla correcta
        cursor.execute("""
            INSERT INTO respuestas_individual (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE opcion_id = %s, es_correcta = %s, puntos = %s, tiempo_respuesta = %s
        """, (sesion_id, g.user['id'], question_id, answer_id, es_correcta, puntos, tiempo_respuesta,
              answer_id, es_correcta, puntos, tiempo_respuesta))

        db.commit()

        # Actualizar pregunta actual del usuario
        cursor.execute("""
            UPDATE usuario_estado_individual
            SET pregunta_actual = %s
            WHERE sesion_id = %s AND user_id = %s
        """, (question_id, sesion_id, g.user['id']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({
            'success': True,
            'correct': bool(es_correcta),
            'points': puntos
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/individual/finalizar-sesion', methods=['POST'])
@jwt_required()
def api_finalizar_sesion_individual():
    """API para marcar sesión individual como finalizada y asignar recompensas"""
    if not g.user:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    data = request.get_json()
    sesion_id = data.get('sesion_id')

    db = None
    cursor = None
    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        cursor.execute("""
            UPDATE sesiones_individual
            SET estado = 'finalizado', finished_at = NOW()
            WHERE id = %s
        """, (sesion_id,))

        db.commit()

        # Asignar puntosmoneda a los ganadores
        asignar_puntosmoneda_individual(sesion_id)

        return jsonify({'success': True})

    except Exception as e:
        if db:
            db.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/individual/resultados/<int:sesion_id>')
def individual_resultados(sesion_id):
    """Página de resultados finales de la sesión individual"""
    if not g.user:
        flash('Debes iniciar sesión para ver los resultados', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión incluyendo el iniciador
        cursor.execute("""
            SELECT si.id, si.cuestionario_id, si.iniciado_por, c.titulo as cuestionario_titulo
            FROM sesiones_individual si
            JOIN cuestionarios c ON si.cuestionario_id = c.id
            WHERE si.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('my_quizzes'))

        # Verificar si el usuario actual es el creador/anfitrión
        es_creador = (g.user['id'] == sesion['iniciado_por'])

        # Obtener el total de preguntas
        cursor.execute("""
            SELECT COUNT(p.id) as total_preguntas
            FROM preguntas p
            WHERE p.cuestionario_id = %s
        """, (sesion['cuestionario_id'],))
        total_preguntas = cursor.fetchone()['total_preguntas']

        # Obtener resultados de todos los participantes
        cursor.execute("""
            SELECT
                u.id,
                u.username,
                COALESCE(SUM(ri.puntos), 0) as score,
                COALESCE(SUM(ri.es_correcta), 0) as correct_answers,
                COUNT(DISTINCT ri.pregunta_id) as preguntas_respondidas
            FROM usuario_estado_individual uei
            JOIN users u ON uei.user_id = u.id
            LEFT JOIN respuestas_individual ri ON ri.user_id = u.id AND ri.sesion_id = %s
            WHERE uei.sesion_id = %s
            GROUP BY u.id, u.username
            ORDER BY score DESC, correct_answers DESC
        """, (sesion_id, sesion_id))
        resultados = cursor.fetchall()

        cursor.close()
        db.close()

        return render_template('individual_resultados.html',
                             sesion_id=sesion_id,
                             cuestionario_titulo=sesion['cuestionario_titulo'],
                             total_preguntas=total_preguntas,
                             resultados=resultados,
                             es_creador=es_creador)

    except Exception as e:
        flash(f'Error al cargar resultados: {str(e)}', 'error')
        return redirect(url_for('my_quizzes'))

# =====================
#  Exportar Resultados a Excel
# =====================

@app.route('/grupo/resultados/<int:sesion_id>/exportar-excel')
def exportar_resultados_grupo_excel(sesion_id):
    """Exportar resultados de grupo a Excel"""
    if not g.user:
        flash('Debes iniciar sesión', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión
        cursor.execute("""
            SELECT sg.id, sg.grupo_id, sg.cuestionario_id,
                   g.nombre as grupo_nombre, c.titulo as cuestionario_titulo
            FROM sesiones_grupo sg
            JOIN grupos g ON sg.grupo_id = g.id
            JOIN cuestionarios c ON sg.cuestionario_id = c.id
            WHERE sg.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('grupos'))

        # Obtener el total de preguntas
        cursor.execute("""
            SELECT COUNT(p.id) as total_preguntas
            FROM preguntas p
            WHERE p.cuestionario_id = %s
        """, (sesion['cuestionario_id'],))
        total_preguntas = cursor.fetchone()['total_preguntas']

        # Obtener resultados de todos los participantes
        cursor.execute("""
            SELECT
                u.username,
                COALESCE(SUM(rg.puntos), 0) as score,
                COALESCE(SUM(rg.es_correcta), 0) as correct_answers,
                COUNT(DISTINCT rg.pregunta_id) as preguntas_respondidas
            FROM usuario_estado_grupo ues
            JOIN users u ON ues.user_id = u.id
            LEFT JOIN respuestas_grupo rg ON rg.user_id = u.id AND rg.sesion_id = %s
            WHERE ues.sesion_id = %s
            GROUP BY u.id, u.username
            ORDER BY score DESC, correct_answers DESC
        """, (sesion_id, sesion_id))
        resultados = cursor.fetchall()

        cursor.close()
        db.close()

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Resultados - {sesion['cuestionario_titulo']}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:E2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Grupo: {sesion['grupo_nombre']}"
        subtitle_cell.font = Font(size=12)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:E3')
        date_cell = ws['A3']
        date_cell.value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Posición', 'Jugador', 'Puntaje', 'Respuestas Correctas', 'Preguntas Respondidas']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        for idx, resultado in enumerate(resultados, start=1):
            row = idx + 5
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=resultado['username']).border = border
            ws.cell(row=row, column=3, value=resultado['score']).border = border
            ws.cell(row=row, column=4, value=f"{resultado['correct_answers']}/{total_preguntas}").border = border
            ws.cell(row=row, column=5, value=resultado['preguntas_respondidas']).border = border

            # Alineación
            for col in range(1, 6):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 25

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Resultados_{sesion['grupo_nombre']}_{sesion['cuestionario_titulo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # ✅ SUBIR AUTOMÁTICAMENTE A GOOGLE DRIVE
        output_copy = BytesIO(output.getvalue())  # Copia para Drive
        drive_result = subir_a_google_drive(output_copy, filename)

        if drive_result['success']:
            flash(f'✅ Resultados guardados en Google Drive', 'success')
        elif drive_result.get('needs_auth'):
            flash(f'⚠️ Para guardar en Drive, necesitas autorizar primero.', 'warning')
            session['needs_drive_auth'] = True
        else:
            flash(f'⚠️ Descarga lista, pero hubo un error al guardar en Drive: {drive_result.get("error", "Error desconocido")}', 'warning')

        # Resetear el puntero para la descarga
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error al exportar resultados: {str(e)}', 'error')
        return redirect(url_for('grupo_resultados', sesion_id=sesion_id))

@app.route('/individual/resultados/<int:sesion_id>/exportar-excel')
def exportar_resultados_individual_excel(sesion_id):
    """Exportar resultados individuales a Excel"""
    if not g.user:
        flash('Debes iniciar sesión', 'warning')
        return redirect(url_for('login'))

    try:
        db = bd.obtener_conexion()
        cursor = db.cursor()

        # Obtener información de la sesión
        cursor.execute("""
            SELECT si.id, si.cuestionario_id, c.titulo as cuestionario_titulo
            FROM sesiones_individual si
            JOIN cuestionarios c ON si.cuestionario_id = c.id
            WHERE si.id = %s
        """, (sesion_id,))

        sesion = cursor.fetchone()

        if not sesion:
            flash('Sesión no encontrada', 'error')
            return redirect(url_for('my_quizzes'))

        # Obtener el total de preguntas
        cursor.execute("""
            SELECT COUNT(p.id) as total_preguntas
            FROM preguntas p
            WHERE p.cuestionario_id = %s
        """, (sesion['cuestionario_id'],))
        total_preguntas = cursor.fetchone()['total_preguntas']

        # Obtener resultados de todos los participantes
        cursor.execute("""
            SELECT
                u.username,
                COALESCE(SUM(ri.puntos), 0) as score,
                COALESCE(SUM(ri.es_correcta), 0) as correct_answers,
                COUNT(DISTINCT ri.pregunta_id) as preguntas_respondidas
            FROM usuario_estado_individual uei
            JOIN users u ON uei.user_id = u.id
            LEFT JOIN respuestas_individual ri ON ri.user_id = u.id AND ri.sesion_id = %s
            WHERE uei.sesion_id = %s
            GROUP BY u.id, u.username
            ORDER BY score DESC, correct_answers DESC
        """, (sesion_id, sesion_id))
        resultados = cursor.fetchall()

        cursor.close()
        db.close()

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Resultados - {sesion['cuestionario_titulo']}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:E2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = "Sesión Individual"
        subtitle_cell.font = Font(size=12)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:E3')
        date_cell = ws['A3']
        date_cell.value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Posición', 'Jugador', 'Puntaje', 'Respuestas Correctas', 'Preguntas Respondidas']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        for idx, resultado in enumerate(resultados, start=1):
            row = idx + 5
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=resultado['username']).border = border
            ws.cell(row=row, column=3, value=resultado['score']).border = border
            ws.cell(row=row, column=4, value=f"{resultado['correct_answers']}/{total_preguntas}").border = border
            ws.cell(row=row, column=5, value=resultado['preguntas_respondidas']).border = border

            # Alineación
            for col in range(1, 6):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 25

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Resultados_{sesion['cuestionario_titulo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # ✅ SUBIR AUTOMÁTICAMENTE A GOOGLE DRIVE
        output_copy = BytesIO(output.getvalue())  # Copia para Drive
        drive_result = subir_a_google_drive(output_copy, filename)

        if drive_result['success']:
            flash(f'✅ Resultados guardados en Google Drive', 'success')
        elif drive_result.get('needs_auth'):
            flash(f'⚠️ Para guardar en Drive, necesitas autorizar primero.', 'warning')
            session['needs_drive_auth'] = True
        else:
            flash(f'⚠️ Descarga lista, pero hubo un error al guardar en Drive: {drive_result.get("error", "Error desconocido")}', 'warning')

        # Resetear el puntero para la descarga
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error al exportar resultados: {str(e)}', 'error')
        return redirect(url_for('individual_resultados', sesion_id=sesion_id))

# =====================
#  Descarga de Plantilla de Ejemplo
# =====================
@app.route('/descargar-plantilla-ejemplo')
def descargar_plantilla_ejemplo():
    """Descargar la plantilla de ejemplo de preguntas en Excel"""
    try:
        import os
        # La plantilla debe estar en la raíz del proyecto Kahoot
        plantilla_path = os.path.join(os.path.dirname(__file__), 'plantilla_preguntas_ejemplo.xlsx')

        if not os.path.exists(plantilla_path):
            flash('La plantilla de ejemplo no está disponible', 'error')
            return redirect(url_for('editor'))

        return send_file(
            plantilla_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_preguntas_ejemplo.xlsx'
        )
    except Exception as e:
        flash(f'Error al descargar la plantilla: {str(e)}', 'error')
        return redirect(url_for('editor'))

#----------------------------------------



# =====================================================
# HELPER FUNCTIONS PARA APIs
# =====================================================

def respuesta_exitosa(data, mensaje="Operación exitosa"):
    """Formato estándar de respuesta exitosa"""
    return jsonify({
        "code": 1,
        "data": data,
        "message": mensaje
    })


def respuesta_error(mensaje, data=None):
    """Formato estándar de respuesta de error"""
    return jsonify({
        "code": 0,
        "data": data if data else {},
        "message": mensaje
    })


def obtener_datos_request():
    """
    Devuelve un diccionario con los datos enviados en la petición
    sin importar si llegaron como JSON o form-data.
    """
    if request.is_json:
        data = request.get_json(silent=True)
        if isinstance(data, dict):
            return data
    if request.form:
        return request.form.to_dict(flat=True)
    return {}


def obtener_id_param(datos, nombre="id"):
    """
    Busca un identificador en el cuerpo de la petición o en los query params.
    """
    valor = datos.get(nombre)
    if valor is None or valor == "":
        valor = request.args.get(nombre)
    return valor


def validar_entero(valor, nombre="id"):
    """
    Valida que un valor sea un entero y devuelve su versión normalizada.
    """
    if valor is None or str(valor).strip() == "":
        raise ValueError(f"Debe proporcionar el campo '{nombre}'")
    try:
        return int(valor)
    except (TypeError, ValueError):
        raise ValueError(f"El campo '{nombre}' debe ser numérico")


# =====================================================
# APIs TABLA: users
# =====================================================

@app.route("/api/users/registrar", methods=['POST'])
@jwt_required()
def api_registrar_user():
    """Registra un nuevo usuario (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        username = datos.get("username")
        email = datos.get("email")
        password = datos.get("password")
        role = datos.get("role", "usuario")
        puntosmoneda = datos.get("puntosmoneda", 0)

        # Encriptar password antes de guardar
        password_encriptado = encriptar_sha256(password)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO users (username, email, password, role, puntosmoneda) VALUES (%s, %s, %s, %s, %s)",
                    (username, email, password_encriptado, role, puntosmoneda)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Usuario registrado correctamente")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return respuesta_error(f"Error al registrar usuario: {repr(e)}")


@app.route("/api/users/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_user():
    """Actualiza un usuario existente (Requiere JWT)"""
    try:
        datos = request.json
        id_user = datos.get("id")
        username = datos.get("username")
        email = datos.get("email")
        password = datos.get("password")
        role = datos.get("role")
        puntosmoneda = datos.get("puntosmoneda")

        # Encriptar password si se proporciona
        if password:
            password = encriptar_sha256(password)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "UPDATE users SET username=%s, email=%s, password=%s, role=%s, puntosmoneda=%s WHERE id=%s",
                    (username, email, password, role, puntosmoneda, id_user)
                )
                conexion.commit()

        return respuesta_exitosa({"id": id_user}, "Usuario actualizado correctamente")
    except Exception as e:
        import sys
        import traceback
        print("\n=== ERROR EN ACTUALIZAR USER ===", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.stderr.flush()
        return jsonify({
            "code": 0,
            "data": {},
            "message": "Error al actualizar usuario"
        }), 400


@app.route("/api/users/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_user_id(id=0):
    """Obtiene un usuario por ID"""
    try:
        if request.method == 'POST':
            id = request.json.get("id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM users WHERE id=%s", (id,))
                usuario = cursor.fetchone()

        if usuario:
            return respuesta_exitosa(usuario, "Usuario obtenido correctamente")
        else:
            return respuesta_error("Usuario no encontrado"), 401
    except Exception as e:
        return respuesta_error(f"Error al obtener usuario: {repr(e)}")


@app.route("/api/users/obtener", methods=['GET'])
@jwt_required()
def api_obtener_users():
    """Obtiene todos los usuarios"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM users")
                usuarios = cursor.fetchall()

        return respuesta_exitosa(usuarios, f"Se encontraron {len(usuarios)} usuarios")
    except Exception as e:
        return respuesta_error(f"Error al obtener usuarios: {repr(e)}")


@app.route("/api/users/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_user():
    """Elimina un usuario (Requiere JWT)"""
    try:
        datos = {}
        if request.is_json:
            datos = request.get_json(silent=True) or {}
        elif request.form:
            datos = request.form

        id_user = datos.get("id") or request.args.get("id")

        if not id_user:
            return respuesta_error("Debe proporcionar el id del usuario a eliminar")

        try:
            id_user = int(id_user)
        except (TypeError, ValueError):
            return respuesta_error("El id del usuario debe ser numérico")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM users WHERE id=%s", (id_user,))
                if filas == 0:
                    return respuesta_error("Usuario no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_user}, "Usuario eliminado correctamente")
    except Exception as e:
        import sys
        import traceback
        print("\n=== ERROR EN ELIMINAR USER ===", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.stderr.flush()
        return jsonify({
            "code": 0,
            "data": {},
            "message": "Error al eliminar usuario"
        })


# =====================================================
# APIs TABLA: grupos
# =====================================================

@app.route("/api/grupos/registrar", methods=['POST'])
@jwt_required()
def api_registrar_grupo():
    """Registra un nuevo grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        nombre = datos.get("nombre")
        descripcion = datos.get("descripcion")
        codigo = datos.get("codigo")
        es_publico = datos.get("es_publico", 0)
        admin_id = datos.get("admin_id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO grupos (nombre, descripcion, codigo, es_publico, admin_id) VALUES (%s, %s, %s, %s, %s)",
                    (nombre, descripcion, codigo, es_publico, admin_id)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Grupo registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar grupo: {str(e)}")


@app.route("/api/grupos/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_grupo():
    """Actualiza un grupo existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_grupo = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        nombre = datos.get("nombre")
        descripcion = datos.get("descripcion")
        codigo = datos.get("codigo")
        es_publico = datos.get("es_publico")
        admin_id = datos.get("admin_id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE grupos SET nombre=%s, descripcion=%s, codigo=%s, es_publico=%s, admin_id=%s WHERE id=%s",
                    (nombre, descripcion, codigo, es_publico, admin_id, id_grupo)
                )
                if filas == 0:
                    return respuesta_error("Grupo no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_grupo}, "Grupo actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar grupo: {str(e)}")


@app.route("/api/grupos/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_grupo_id(id=0):
    """Obtiene un grupo por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM grupos WHERE id=%s", (target_id,))
                grupo = cursor.fetchone()

        if grupo:
            return respuesta_exitosa(grupo, "Grupo obtenido correctamente")
        else:
            return respuesta_error("Grupo no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener grupo: {str(e)}")


@app.route("/api/grupos/obtener", methods=['GET'])
@jwt_required()
def api_obtener_grupos():
    """Obtiene todos los grupos"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM grupos")
                grupos = cursor.fetchall()

        return respuesta_exitosa(grupos, f"Se encontraron {len(grupos)} grupos")
    except Exception as e:
        return respuesta_error(f"Error al obtener grupos: {str(e)}")


@app.route("/api/grupos/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_grupo():
    """Elimina un grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_grupo = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM grupos WHERE id=%s", (id_grupo,))
                if filas == 0:
                    return respuesta_error("Grupo no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_grupo}, "Grupo eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar grupo: {str(e)}")


# =====================================================
# APIs TABLA: grupo_miembros
# =====================================================

@app.route("/api/grupo_miembros/registrar", methods=['POST'])
@jwt_required()
def api_registrar_grupo_miembro():
    """Registra un nuevo miembro en un grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        grupo_id = datos.get("grupo_id")
        user_id = datos.get("user_id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO grupo_miembros (grupo_id, user_id) VALUES (%s, %s)",
                    (grupo_id, user_id)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Miembro agregado al grupo correctamente")
    except Exception as e:
        return respuesta_error(f"Error al agregar miembro: {str(e)}")


@app.route("/api/grupo_miembros/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_grupo_miembro():
    """Actualiza un miembro de grupo existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_miembro = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        grupo_id = datos.get("grupo_id")
        user_id = datos.get("user_id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE grupo_miembros SET grupo_id=%s, user_id=%s WHERE id=%s",
                    (grupo_id, user_id, id_miembro)
                )
                if filas == 0:
                    return respuesta_error("Miembro no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_miembro}, "Miembro actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar miembro: {str(e)}")


@app.route("/api/grupo_miembros/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_grupo_miembro_id(id=0):
    """Obtiene un miembro de grupo por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM grupo_miembros WHERE id=%s", (target_id,))
                miembro = cursor.fetchone()

        if miembro:
            return respuesta_exitosa(miembro, "Miembro obtenido correctamente")
        else:
            return respuesta_error("Miembro no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener miembro: {str(e)}")


@app.route("/api/grupo_miembros/obtener", methods=['GET'])
@jwt_required()
def api_obtener_grupo_miembros():
    """Obtiene todos los miembros de grupos"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM grupo_miembros")
                miembros = cursor.fetchall()

        return respuesta_exitosa(miembros, f"Se encontraron {len(miembros)} miembros")
    except Exception as e:
        return respuesta_error(f"Error al obtener miembros: {str(e)}")


@app.route("/api/grupo_miembros/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_grupo_miembro():
    """Elimina un miembro de un grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_miembro = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM grupo_miembros WHERE id=%s", (id_miembro,))
                if filas == 0:
                    return respuesta_error("Miembro no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_miembro}, "Miembro eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar miembro: {str(e)}")


# =====================================================
# APIs TABLA: cuestionarios
# =====================================================

@app.route("/api/cuestionarios/registrar", methods=['POST'])
@jwt_required()
def api_registrar_cuestionario():
    """Registra un nuevo cuestionario (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        user_id = datos.get("user_id")
        titulo = datos.get("titulo")
        descripcion = datos.get("descripcion")
        imagen_portada = datos.get("imagen_portada")
        pin = datos.get("pin")
        estado = datos.get("estado", "publico")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO cuestionarios (user_id, titulo, descripcion, imagen_portada, pin, estado) VALUES (%s, %s, %s, %s, %s, %s)",
                    (user_id, titulo, descripcion, imagen_portada, pin, estado)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Cuestionario registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar cuestionario: {str(e)}")


@app.route("/api/cuestionarios/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_cuestionario():
    """Actualiza un cuestionario existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_cuestionario = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        user_id = datos.get("user_id")
        titulo = datos.get("titulo")
        descripcion = datos.get("descripcion")
        imagen_portada = datos.get("imagen_portada")
        pin = datos.get("pin")
        estado = datos.get("estado")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE cuestionarios SET user_id=%s, titulo=%s, descripcion=%s, imagen_portada=%s, pin=%s, estado=%s WHERE id=%s",
                    (user_id, titulo, descripcion, imagen_portada, pin, estado, id_cuestionario)
                )
                if filas == 0:
                    return respuesta_error("Cuestionario no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_cuestionario}, "Cuestionario actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar cuestionario: {str(e)}")


@app.route("/api/cuestionarios/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_cuestionario_id(id=0):
    """Obtiene un cuestionario por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM cuestionarios WHERE id=%s", (target_id,))
                cuestionario = cursor.fetchone()

        if cuestionario:
            return respuesta_exitosa(cuestionario, "Cuestionario obtenido correctamente")
        else:
            return respuesta_error("Cuestionario no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener cuestionario: {str(e)}")


@app.route("/api/cuestionarios/obtener", methods=['GET'])
@jwt_required()
def api_obtener_cuestionarios():
    """Obtiene todos los cuestionarios"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM cuestionarios")
                cuestionarios = cursor.fetchall()

        return respuesta_exitosa(cuestionarios, f"Se encontraron {len(cuestionarios)} cuestionarios")
    except Exception as e:
        return respuesta_error(f"Error al obtener cuestionarios: {str(e)}")


@app.route("/api/cuestionarios/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_cuestionario():
    """Elimina un cuestionario (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_cuestionario = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM cuestionarios WHERE id=%s", (id_cuestionario,))
                if filas == 0:
                    return respuesta_error("Cuestionario no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_cuestionario}, "Cuestionario eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar cuestionario: {str(e)}")


# =====================================================
# APIs TABLA: preguntas
# =====================================================

@app.route("/api/preguntas/registrar", methods=['POST'])
@jwt_required()
def api_registrar_pregunta():
    """Registra una nueva pregunta (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        cuestionario_id = datos.get("cuestionario_id")
        tipo_pregunta = datos.get("tipo_pregunta")
        texto_pregunta = datos.get("texto_pregunta")
        imagen_pregunta = datos.get("imagen_pregunta")
        orden = datos.get("orden", 0)
        tiempo_limite = datos.get("tiempo_limite", 30)
        puntos = datos.get("puntos", 1)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO preguntas (cuestionario_id, tipo_pregunta, texto_pregunta, imagen_pregunta, orden, tiempo_limite, puntos) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (cuestionario_id, tipo_pregunta, texto_pregunta, imagen_pregunta, orden, tiempo_limite, puntos)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Pregunta registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar pregunta: {str(e)}")


@app.route("/api/preguntas/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_pregunta():
    """Actualiza una pregunta existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_pregunta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        cuestionario_id = datos.get("cuestionario_id")
        tipo_pregunta = datos.get("tipo_pregunta")
        texto_pregunta = datos.get("texto_pregunta")
        imagen_pregunta = datos.get("imagen_pregunta")
        orden = datos.get("orden")
        tiempo_limite = datos.get("tiempo_limite")
        puntos = datos.get("puntos")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE preguntas SET cuestionario_id=%s, tipo_pregunta=%s, texto_pregunta=%s, imagen_pregunta=%s, orden=%s, tiempo_limite=%s, puntos=%s WHERE id=%s",
                    (cuestionario_id, tipo_pregunta, texto_pregunta, imagen_pregunta, orden, tiempo_limite, puntos, id_pregunta)
                )
                if filas == 0:
                    return respuesta_error("Pregunta no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_pregunta}, "Pregunta actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar pregunta: {str(e)}")


@app.route("/api/preguntas/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_pregunta_id(id=0):
    """Obtiene una pregunta por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM preguntas WHERE id=%s", (target_id,))
                pregunta = cursor.fetchone()

        if pregunta:
            return respuesta_exitosa(pregunta, "Pregunta obtenida correctamente")
        else:
            return respuesta_error("Pregunta no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener pregunta: {str(e)}")


@app.route("/api/preguntas/obtener", methods=['GET'])
@jwt_required()
def api_obtener_preguntas():
    """Obtiene todas las preguntas"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM preguntas")
                preguntas = cursor.fetchall()

        return respuesta_exitosa(preguntas, f"Se encontraron {len(preguntas)} preguntas")
    except Exception as e:
        return respuesta_error(f"Error al obtener preguntas: {str(e)}")


@app.route("/api/preguntas/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_pregunta():
    """Elimina una pregunta (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_pregunta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM preguntas WHERE id=%s", (id_pregunta,))
                if filas == 0:
                    return respuesta_error("Pregunta no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_pregunta}, "Pregunta eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar pregunta: {str(e)}")


# =====================================================
# APIs TABLA: opciones_respuesta
# =====================================================

@app.route("/api/opciones_respuesta/registrar", methods=['POST'])
@jwt_required()
def api_registrar_opcion_respuesta():
    """Registra una nueva opción de respuesta (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        pregunta_id = datos.get("pregunta_id")
        texto_opcion = datos.get("texto_opcion")
        es_correcta = datos.get("es_correcta", 0)
        orden = datos.get("orden", 0)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO opciones_respuesta (pregunta_id, texto_opcion, es_correcta, orden) VALUES (%s, %s, %s, %s)",
                    (pregunta_id, texto_opcion, es_correcta, orden)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Opción de respuesta registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar opción de respuesta: {str(e)}")


@app.route("/api/opciones_respuesta/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_opcion_respuesta():
    """Actualiza una opción de respuesta existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_opcion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        pregunta_id = datos.get("pregunta_id")
        texto_opcion = datos.get("texto_opcion")
        es_correcta = datos.get("es_correcta")
        orden = datos.get("orden")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE opciones_respuesta SET pregunta_id=%s, texto_opcion=%s, es_correcta=%s, orden=%s WHERE id=%s",
                    (pregunta_id, texto_opcion, es_correcta, orden, id_opcion)
                )
                if filas == 0:
                    return respuesta_error("Opción de respuesta no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_opcion}, "Opción de respuesta actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar opción de respuesta: {str(e)}")


@app.route("/api/opciones_respuesta/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_opcion_respuesta_id(id=0):
    """Obtiene una opción de respuesta por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM opciones_respuesta WHERE id=%s", (target_id,))
                opcion = cursor.fetchone()

        if opcion:
            return respuesta_exitosa(opcion, "Opción de respuesta obtenida correctamente")
        else:
            return respuesta_error("Opción de respuesta no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener opción de respuesta: {str(e)}")


@app.route("/api/opciones_respuesta/obtener", methods=['GET'])
@jwt_required()
def api_obtener_opciones_respuesta():
    """Obtiene todas las opciones de respuesta"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM opciones_respuesta")
                opciones = cursor.fetchall()

        return respuesta_exitosa(opciones, f"Se encontraron {len(opciones)} opciones de respuesta")
    except Exception as e:
        return respuesta_error(f"Error al obtener opciones de respuesta: {str(e)}")


@app.route("/api/opciones_respuesta/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_opcion_respuesta():
    """Elimina una opción de respuesta (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_opcion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM opciones_respuesta WHERE id=%s", (id_opcion,))
                if filas == 0:
                    return respuesta_error("Opción de respuesta no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_opcion}, "Opción de respuesta eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar opción de respuesta: {str(e)}")


# =====================================================
# APIs TABLA: sesiones_grupo
# =====================================================

@app.route("/api/sesiones_grupo/registrar", methods=['POST'])
@jwt_required()
def api_registrar_sesion_grupo():
    """Registra una nueva sesión de grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        grupo_id = datos.get("grupo_id")
        cuestionario_id = datos.get("cuestionario_id")
        iniciado_por = datos.get("iniciado_por")
        session_code = datos.get("session_code")
        estado = datos.get("estado", "esperando")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO sesiones_grupo (grupo_id, cuestionario_id, iniciado_por, session_code, estado) VALUES (%s, %s, %s, %s, %s)",
                    (grupo_id, cuestionario_id, iniciado_por, session_code, estado)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Sesión de grupo registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar sesión de grupo: {str(e)}")


@app.route("/api/sesiones_grupo/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_sesion_grupo():
    """Actualiza una sesión de grupo existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        grupo_id = datos.get("grupo_id")
        cuestionario_id = datos.get("cuestionario_id")
        iniciado_por = datos.get("iniciado_por")
        session_code = datos.get("session_code")
        estado = datos.get("estado")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE sesiones_grupo SET grupo_id=%s, cuestionario_id=%s, iniciado_por=%s, session_code=%s, estado=%s WHERE id=%s",
                    (grupo_id, cuestionario_id, iniciado_por, session_code, estado, id_sesion)
                )
                if filas == 0:
                    return respuesta_error("Sesión de grupo no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión de grupo actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar sesión de grupo: {str(e)}")


@app.route("/api/sesiones_grupo/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_sesion_grupo_id(id=0):
    """Obtiene una sesión de grupo por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_grupo WHERE id=%s", (target_id,))
                sesion = cursor.fetchone()

        if sesion:
            return respuesta_exitosa(sesion, "Sesión de grupo obtenida correctamente")
        else:
            return respuesta_error("Sesión de grupo no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesión de grupo: {str(e)}")


@app.route("/api/sesiones_grupo/obtener", methods=['GET'])
@jwt_required()
def api_obtener_sesiones_grupo():
    """Obtiene todas las sesiones de grupo"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_grupo")
                sesiones = cursor.fetchall()

        return respuesta_exitosa(sesiones, f"Se encontraron {len(sesiones)} sesiones de grupo")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesiones de grupo: {str(e)}")


@app.route("/api/sesiones_grupo/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_sesion_grupo():
    """Elimina una sesión de grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM sesiones_grupo WHERE id=%s", (id_sesion,))
                if filas == 0:
                    return respuesta_error("Sesión de grupo no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión de grupo eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar sesión de grupo: {str(e)}")


# =====================================================
# APIs TABLA: sesiones_individual
# =====================================================

@app.route("/api/sesiones_individual/registrar", methods=['POST'])
@jwt_required()
def api_registrar_sesion_individual():
    """Registra una nueva sesión individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        cuestionario_id = datos.get("cuestionario_id")
        iniciado_por = datos.get("iniciado_por")
        session_code = datos.get("session_code")
        estado = datos.get("estado", "esperando")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO sesiones_individual (cuestionario_id, iniciado_por, session_code, estado) VALUES (%s, %s, %s, %s)",
                    (cuestionario_id, iniciado_por, session_code, estado)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Sesión individual registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar sesión individual: {str(e)}")


@app.route("/api/sesiones_individual/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_sesion_individual():
    """Actualiza una sesión individual existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        cuestionario_id = datos.get("cuestionario_id")
        iniciado_por = datos.get("iniciado_por")
        session_code = datos.get("session_code")
        estado = datos.get("estado")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE sesiones_individual SET cuestionario_id=%s, iniciado_por=%s, session_code=%s, estado=%s WHERE id=%s",
                    (cuestionario_id, iniciado_por, session_code, estado, id_sesion)
                )
                if filas == 0:
                    return respuesta_error("Sesión individual no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión individual actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar sesión individual: {str(e)}")


@app.route("/api/sesiones_individual/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_sesion_individual_id(id=0):
    """Obtiene una sesión individual por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_individual WHERE id=%s", (target_id,))
                sesion = cursor.fetchone()

        if sesion:
            return respuesta_exitosa(sesion, "Sesión individual obtenida correctamente")
        else:
            return respuesta_error("Sesión individual no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesión individual: {str(e)}")


@app.route("/api/sesiones_individual/obtener", methods=['GET'])
@jwt_required()
def api_obtener_sesiones_individual():
    """Obtiene todas las sesiones individuales"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_individual")
                sesiones = cursor.fetchall()

        return respuesta_exitosa(sesiones, f"Se encontraron {len(sesiones)} sesiones individuales")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesiones individuales: {str(e)}")


@app.route("/api/sesiones_individual/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_sesion_individual():
    """Elimina una sesión individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM sesiones_individual WHERE id=%s", (id_sesion,))
                if filas == 0:
                    return respuesta_error("Sesión individual no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión individual eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar sesión individual: {str(e)}")


# =====================================================
# APIs TABLA: sesiones_juego
# =====================================================

@app.route("/api/sesiones_juego/registrar", methods=['POST'])
@jwt_required()
def api_registrar_sesion_juego():
    """Registra una nueva sesión de juego (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        cuestionario_id = datos.get("cuestionario_id")
        user_id = datos.get("user_id")
        created_by = datos.get("created_by")
        estado = datos.get("estado", "esperando")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO sesiones_juego (cuestionario_id, user_id, created_by, estado) VALUES (%s, %s, %s, %s)",
                    (cuestionario_id, user_id, created_by, estado)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Sesión de juego registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar sesión de juego: {str(e)}")


@app.route("/api/sesiones_juego/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_sesion_juego():
    """Actualiza una sesión de juego existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        cuestionario_id = datos.get("cuestionario_id")
        user_id = datos.get("user_id")
        created_by = datos.get("created_by")
        estado = datos.get("estado")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE sesiones_juego SET cuestionario_id=%s, user_id=%s, created_by=%s, estado=%s WHERE id=%s",
                    (cuestionario_id, user_id, created_by, estado, id_sesion)
                )
                if filas == 0:
                    return respuesta_error("Sesión de juego no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión de juego actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar sesión de juego: {str(e)}")


@app.route("/api/sesiones_juego/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_sesion_juego_id(id=0):
    """Obtiene una sesión de juego por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_juego WHERE id=%s", (target_id,))
                sesion = cursor.fetchone()

        if sesion:
            return respuesta_exitosa(sesion, "Sesión de juego obtenida correctamente")
        else:
            return respuesta_error("Sesión de juego no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesión de juego: {str(e)}")


@app.route("/api/sesiones_juego/obtener", methods=['GET'])
@jwt_required()
def api_obtener_sesiones_juego():
    """Obtiene todas las sesiones de juego"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM sesiones_juego")
                sesiones = cursor.fetchall()

        return respuesta_exitosa(sesiones, f"Se encontraron {len(sesiones)} sesiones de juego")
    except Exception as e:
        return respuesta_error(f"Error al obtener sesiones de juego: {str(e)}")


@app.route("/api/sesiones_juego/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_sesion_juego():
    """Elimina una sesión de juego (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_sesion = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM sesiones_juego WHERE id=%s", (id_sesion,))
                if filas == 0:
                    return respuesta_error("Sesión de juego no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_sesion}, "Sesión de juego eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar sesión de juego: {str(e)}")


# =====================================================
# APIs TABLA: respuestas_grupo
# =====================================================

@app.route("/api/respuestas_grupo/registrar", methods=['POST'])
@jwt_required()
def api_registrar_respuesta_grupo():
    """Registra una nueva respuesta de grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta", 0)
        puntos = datos.get("puntos", 0)
        tiempo_respuesta = datos.get("tiempo_respuesta")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO respuestas_grupo (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Respuesta de grupo registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar respuesta de grupo: {str(e)}")


@app.route("/api/respuestas_grupo/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_respuesta_grupo():
    """Actualiza una respuesta de grupo existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta")
        puntos = datos.get("puntos")
        tiempo_respuesta = datos.get("tiempo_respuesta")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE respuestas_grupo SET sesion_id=%s, user_id=%s, pregunta_id=%s, opcion_id=%s, es_correcta=%s, puntos=%s, tiempo_respuesta=%s WHERE id=%s",
                    (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta, id_respuesta)
                )
                if filas == 0:
                    return respuesta_error("Respuesta de grupo no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta de grupo actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar respuesta de grupo: {str(e)}")


@app.route("/api/respuestas_grupo/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_respuesta_grupo_id(id=0):
    """Obtiene una respuesta de grupo por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_grupo WHERE id=%s", (target_id,))
                respuesta = cursor.fetchone()

        if respuesta:
            return respuesta_exitosa(respuesta, "Respuesta de grupo obtenida correctamente")
        else:
            return respuesta_error("Respuesta de grupo no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuesta de grupo: {str(e)}")


@app.route("/api/respuestas_grupo/obtener", methods=['GET'])
@jwt_required()
def api_obtener_respuestas_grupo():
    """Obtiene todas las respuestas de grupo"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_grupo")
                respuestas = cursor.fetchall()

        return respuesta_exitosa(respuestas, f"Se encontraron {len(respuestas)} respuestas de grupo")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuestas de grupo: {str(e)}")


@app.route("/api/respuestas_grupo/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_respuesta_grupo():
    """Elimina una respuesta de grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM respuestas_grupo WHERE id=%s", (id_respuesta,))
                if filas == 0:
                    return respuesta_error("Respuesta de grupo no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta de grupo eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar respuesta de grupo: {str(e)}")


# =====================================================
# APIs TABLA: respuestas_individual
# =====================================================

@app.route("/api/respuestas_individual/registrar", methods=['POST'])
@jwt_required()
def api_registrar_respuesta_individual():
    """Registra una nueva respuesta individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta", 0)
        puntos = datos.get("puntos", 0)
        tiempo_respuesta = datos.get("tiempo_respuesta")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO respuestas_individual (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Respuesta individual registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar respuesta individual: {str(e)}")


@app.route("/api/respuestas_individual/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_respuesta_individual():
    """Actualiza una respuesta individual existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta")
        puntos = datos.get("puntos")
        tiempo_respuesta = datos.get("tiempo_respuesta")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE respuestas_individual SET sesion_id=%s, user_id=%s, pregunta_id=%s, opcion_id=%s, es_correcta=%s, puntos=%s, tiempo_respuesta=%s WHERE id=%s",
                    (sesion_id, user_id, pregunta_id, opcion_id, es_correcta, puntos, tiempo_respuesta, id_respuesta)
                )
                if filas == 0:
                    return respuesta_error("Respuesta individual no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta individual actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar respuesta individual: {str(e)}")


@app.route("/api/respuestas_individual/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_respuesta_individual_id(id=0):
    """Obtiene una respuesta individual por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_individual WHERE id=%s", (target_id,))
                respuesta = cursor.fetchone()

        if respuesta:
            return respuesta_exitosa(respuesta, "Respuesta individual obtenida correctamente")
        else:
            return respuesta_error("Respuesta individual no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuesta individual: {str(e)}")


@app.route("/api/respuestas_individual/obtener", methods=['GET'])
@jwt_required()
def api_obtener_respuestas_individual():
    """Obtiene todas las respuestas individuales"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_individual")
                respuestas = cursor.fetchall()

        return respuesta_exitosa(respuestas, f"Se encontraron {len(respuestas)} respuestas individuales")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuestas individuales: {str(e)}")


@app.route("/api/respuestas_individual/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_respuesta_individual():
    """Elimina una respuesta individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM respuestas_individual WHERE id=%s", (id_respuesta,))
                if filas == 0:
                    return respuesta_error("Respuesta individual no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta individual eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar respuesta individual: {str(e)}")


# =====================================================
# APIs TABLA: respuestas_participantes
# =====================================================

@app.route("/api/respuestas_participantes/registrar", methods=['POST'])
@jwt_required()
def api_registrar_respuesta_participante():
    """Registra una nueva respuesta de participante (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        sesion_juego_id = datos.get("sesion_juego_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta", 0)
        tiempo_respuesta = datos.get("tiempo_respuesta")
        puntos_obtenidos = datos.get("puntos_obtenidos", 0)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO respuestas_participantes (sesion_juego_id, user_id, pregunta_id, opcion_id, es_correcta, tiempo_respuesta, puntos_obtenidos) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (sesion_juego_id, user_id, pregunta_id, opcion_id, es_correcta, tiempo_respuesta, puntos_obtenidos)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Respuesta de participante registrada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar respuesta de participante: {str(e)}")


@app.route("/api/respuestas_participantes/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_respuesta_participante():
    """Actualiza una respuesta de participante existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        sesion_juego_id = datos.get("sesion_juego_id")
        user_id = datos.get("user_id")
        pregunta_id = datos.get("pregunta_id")
        opcion_id = datos.get("opcion_id")
        es_correcta = datos.get("es_correcta")
        tiempo_respuesta = datos.get("tiempo_respuesta")
        puntos_obtenidos = datos.get("puntos_obtenidos")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE respuestas_participantes SET sesion_juego_id=%s, user_id=%s, pregunta_id=%s, opcion_id=%s, es_correcta=%s, tiempo_respuesta=%s, puntos_obtenidos=%s WHERE id=%s",
                    (sesion_juego_id, user_id, pregunta_id, opcion_id, es_correcta, tiempo_respuesta, puntos_obtenidos, id_respuesta)
                )
                if filas == 0:
                    return respuesta_error("Respuesta de participante no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta de participante actualizada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar respuesta de participante: {str(e)}")


@app.route("/api/respuestas_participantes/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_respuesta_participante_id(id=0):
    """Obtiene una respuesta de participante por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_participantes WHERE id=%s", (target_id,))
                respuesta = cursor.fetchone()

        if respuesta:
            return respuesta_exitosa(respuesta, "Respuesta de participante obtenida correctamente")
        else:
            return respuesta_error("Respuesta de participante no encontrada")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuesta de participante: {str(e)}")


@app.route("/api/respuestas_participantes/obtener", methods=['GET'])
@jwt_required()
def api_obtener_respuestas_participantes():
    """Obtiene todas las respuestas de participantes"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM respuestas_participantes")
                respuestas = cursor.fetchall()

        return respuesta_exitosa(respuestas, f"Se encontraron {len(respuestas)} respuestas de participantes")
    except Exception as e:
        return respuesta_error(f"Error al obtener respuestas de participantes: {str(e)}")


@app.route("/api/respuestas_participantes/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_respuesta_participante():
    """Elimina una respuesta de participante (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_respuesta = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM respuestas_participantes WHERE id=%s", (id_respuesta,))
                if filas == 0:
                    return respuesta_error("Respuesta de participante no encontrada")
                conexion.commit()

        return respuesta_exitosa({"id": id_respuesta}, "Respuesta de participante eliminada correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar respuesta de participante: {str(e)}")


# =====================================================
# APIs TABLA: resultados
# =====================================================

@app.route("/api/resultados/registrar", methods=['POST'])
@jwt_required()
def api_registrar_resultado():
    """Registra un nuevo resultado (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        user_id = datos.get("user_id")
        cuestionario_id = datos.get("cuestionario_id")
        sesion_id = datos.get("sesion_id")
        puntaje_obtenido = datos.get("puntaje_obtenido", 0)
        puntaje_maximo = datos.get("puntaje_maximo")
        total_preguntas = datos.get("total_preguntas")
        respuestas_correctas = datos.get("respuestas_correctas", 0)
        tiempo_total = datos.get("tiempo_total")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO resultados (user_id, cuestionario_id, sesion_id, puntaje_obtenido, puntaje_maximo, total_preguntas, respuestas_correctas, tiempo_total) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    (user_id, cuestionario_id, sesion_id, puntaje_obtenido, puntaje_maximo, total_preguntas, respuestas_correctas, tiempo_total)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Resultado registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar resultado: {str(e)}")


@app.route("/api/resultados/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_resultado():
    """Actualiza un resultado existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_resultado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        user_id = datos.get("user_id")
        cuestionario_id = datos.get("cuestionario_id")
        sesion_id = datos.get("sesion_id")
        puntaje_obtenido = datos.get("puntaje_obtenido")
        puntaje_maximo = datos.get("puntaje_maximo")
        total_preguntas = datos.get("total_preguntas")
        respuestas_correctas = datos.get("respuestas_correctas")
        tiempo_total = datos.get("tiempo_total")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE resultados SET user_id=%s, cuestionario_id=%s, sesion_id=%s, puntaje_obtenido=%s, puntaje_maximo=%s, total_preguntas=%s, respuestas_correctas=%s, tiempo_total=%s WHERE id=%s",
                    (user_id, cuestionario_id, sesion_id, puntaje_obtenido, puntaje_maximo, total_preguntas, respuestas_correctas, tiempo_total, id_resultado)
                )
                if filas == 0:
                    return respuesta_error("Resultado no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_resultado}, "Resultado actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar resultado: {str(e)}")


@app.route("/api/resultados/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_resultado_id(id=0):
    """Obtiene un resultado por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM resultados WHERE id=%s", (target_id,))
                resultado = cursor.fetchone()

        if resultado:
            return respuesta_exitosa(resultado, "Resultado obtenido correctamente")
        else:
            return respuesta_error("Resultado no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener resultado: {str(e)}")


@app.route("/api/resultados/obtener", methods=['GET'])
@jwt_required()
def api_obtener_resultados():
    """Obtiene todos los resultados"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM resultados")
                resultados = cursor.fetchall()

        return respuesta_exitosa(resultados, f"Se encontraron {len(resultados)} resultados")
    except Exception as e:
        return respuesta_error(f"Error al obtener resultados: {str(e)}")


@app.route("/api/resultados/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_resultado():
    """Elimina un resultado (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_resultado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM resultados WHERE id=%s", (id_resultado,))
                if filas == 0:
                    return respuesta_error("Resultado no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_resultado}, "Resultado eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar resultado: {str(e)}")


# =====================================================
# APIs TABLA: usuario_estado_grupo
# =====================================================

@app.route("/api/usuario_estado_grupo/registrar", methods=['POST'])
@jwt_required()
def api_registrar_usuario_estado_grupo():
    """Registra un nuevo estado de usuario en grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        esta_listo = datos.get("esta_listo", 0)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO usuario_estado_grupo (sesion_id, user_id, esta_listo) VALUES (%s, %s, %s)",
                    (sesion_id, user_id, esta_listo)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Estado de usuario en grupo registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar estado de usuario en grupo: {str(e)}")


@app.route("/api/usuario_estado_grupo/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_usuario_estado_grupo():
    """Actualiza un estado de usuario en grupo existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_estado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        esta_listo = datos.get("esta_listo")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE usuario_estado_grupo SET sesion_id=%s, user_id=%s, esta_listo=%s WHERE id=%s",
                    (sesion_id, user_id, esta_listo, id_estado)
                )
                if filas == 0:
                    return respuesta_error("Estado de usuario en grupo no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_estado}, "Estado de usuario en grupo actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar estado de usuario en grupo: {str(e)}")


@app.route("/api/usuario_estado_grupo/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_usuario_estado_grupo_id(id=0):
    """Obtiene un estado de usuario en grupo por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuario_estado_grupo WHERE id=%s", (target_id,))
                estado = cursor.fetchone()

        if estado:
            return respuesta_exitosa(estado, "Estado de usuario en grupo obtenido correctamente")
        else:
            return respuesta_error("Estado de usuario en grupo no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener estado de usuario en grupo: {str(e)}")


@app.route("/api/usuario_estado_grupo/obtener", methods=['GET'])
@jwt_required()
def api_obtener_usuario_estados_grupo():
    """Obtiene todos los estados de usuarios en grupo"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuario_estado_grupo")
                estados = cursor.fetchall()

        return respuesta_exitosa(estados, f"Se encontraron {len(estados)} estados de usuarios en grupo")
    except Exception as e:
        return respuesta_error(f"Error al obtener estados de usuarios en grupo: {str(e)}")


@app.route("/api/usuario_estado_grupo/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_usuario_estado_grupo():
    """Elimina un estado de usuario en grupo (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_estado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM usuario_estado_grupo WHERE id=%s", (id_estado,))
                if filas == 0:
                    return respuesta_error("Estado de usuario en grupo no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_estado}, "Estado de usuario en grupo eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar estado de usuario en grupo: {str(e)}")


# =====================================================
# APIs TABLA: usuario_estado_individual
# =====================================================

@app.route("/api/usuario_estado_individual/registrar", methods=['POST'])
@jwt_required()
def api_registrar_usuario_estado_individual():
    """Registra un nuevo estado de usuario individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        esta_listo = datos.get("esta_listo", 0)

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO usuario_estado_individual (sesion_id, user_id, esta_listo) VALUES (%s, %s, %s)",
                    (sesion_id, user_id, esta_listo)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Estado de usuario individual registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar estado de usuario individual: {str(e)}")


@app.route("/api/usuario_estado_individual/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_usuario_estado_individual():
    """Actualiza un estado de usuario individual existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_estado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        sesion_id = datos.get("sesion_id")
        user_id = datos.get("user_id")
        esta_listo = datos.get("esta_listo")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE usuario_estado_individual SET sesion_id=%s, user_id=%s, esta_listo=%s WHERE id=%s",
                    (sesion_id, user_id, esta_listo, id_estado)
                )
                if filas == 0:
                    return respuesta_error("Estado de usuario individual no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_estado}, "Estado de usuario individual actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar estado de usuario individual: {str(e)}")


@app.route("/api/usuario_estado_individual/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_usuario_estado_individual_id(id=0):
    """Obtiene un estado de usuario individual por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuario_estado_individual WHERE id=%s", (target_id,))
                estado = cursor.fetchone()

        if estado:
            return respuesta_exitosa(estado, "Estado de usuario individual obtenido correctamente")
        else:
            return respuesta_error("Estado de usuario individual no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener estado de usuario individual: {str(e)}")


@app.route("/api/usuario_estado_individual/obtener", methods=['GET'])
@jwt_required()
def api_obtener_usuario_estados_individual():
    """Obtiene todos los estados de usuarios individuales"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuario_estado_individual")
                estados = cursor.fetchall()

        return respuesta_exitosa(estados, f"Se encontraron {len(estados)} estados de usuarios individuales")
    except Exception as e:
        return respuesta_error(f"Error al obtener estados de usuarios individuales: {str(e)}")


@app.route("/api/usuario_estado_individual/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_usuario_estado_individual():
    """Elimina un estado de usuario individual (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_estado = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM usuario_estado_individual WHERE id=%s", (id_estado,))
                if filas == 0:
                    return respuesta_error("Estado de usuario individual no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_estado}, "Estado de usuario individual eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar estado de usuario individual: {str(e)}")


# =====================================================
# APIs TABLA: rangos
# Estructura: id, nombre, puntos_minimos, puntos_maximos, icono, color, orden
# =====================================================

@app.route("/api/rangos/registrar", methods=['POST'])
@jwt_required()
def api_registrar_rango():
    """Registra un nuevo rango (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        nombre = datos.get("nombre")
        puntos_minimos = datos.get("puntos_minimos", 0)
        puntos_maximos = datos.get("puntos_maximos")
        icono = datos.get("icono")
        color = datos.get("color")
        orden = datos.get("orden")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO rangos (nombre, puntos_minimos, puntos_maximos, icono, color, orden) VALUES (%s, %s, %s, %s, %s, %s)",
                    (nombre, puntos_minimos, puntos_maximos, icono, color, orden)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Rango registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar rango: {str(e)}")


@app.route("/api/rangos/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_rango():
    """Actualiza un rango existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_rango = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        nombre = datos.get("nombre")
        puntos_minimos = datos.get("puntos_minimos")
        puntos_maximos = datos.get("puntos_maximos")
        icono = datos.get("icono")
        color = datos.get("color")
        orden = datos.get("orden")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE rangos SET nombre=%s, puntos_minimos=%s, puntos_maximos=%s, icono=%s, color=%s, orden=%s WHERE id=%s",
                    (nombre, puntos_minimos, puntos_maximos, icono, color, orden, id_rango)
                )
                if filas == 0:
                    return respuesta_error("Rango no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_rango}, "Rango actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar rango: {str(e)}")


@app.route("/api/rangos/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_rango_id(id=0):
    """Obtiene un rango por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM rangos WHERE id=%s", (target_id,))
                rango = cursor.fetchone()

        if rango:
            return respuesta_exitosa(rango, "Rango obtenido correctamente")
        else:
            return respuesta_error("Rango no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener rango: {str(e)}")


@app.route("/api/rangos/obtener", methods=['GET'])
@jwt_required()
def api_obtener_rangos():
    """Obtiene todos los rangos"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM rangos")
                rangos = cursor.fetchall()

        return respuesta_exitosa(rangos, f"Se encontraron {len(rangos)} rangos")
    except Exception as e:
        return respuesta_error(f"Error al obtener rangos: {str(e)}")


@app.route("/api/rangos/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_rango():
    """Elimina un rango (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_rango = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM rangos WHERE id=%s", (id_rango,))
                if filas == 0:
                    return respuesta_error("Rango no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_rango}, "Rango eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar rango: {str(e)}")


# =====================================================
# APIs TABLA: historial_recompensas
# Estructura: id, user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales
# =====================================================

@app.route("/api/historial_recompensas/registrar", methods=['POST'])
@jwt_required()
def api_registrar_historial_recompensa():
    """Registra un nuevo historial de recompensa (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        user_id = datos.get("user_id")
        sesion_id = datos.get("sesion_id")
        tipo_sesion = datos.get("tipo_sesion")  # 'grupo' o 'individual'
        posicion = datos.get("posicion")
        puntosmoneda_ganados = datos.get("puntosmoneda_ganados")
        puntos_totales = datos.get("puntos_totales")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO historial_recompensas (user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales) VALUES (%s, %s, %s, %s, %s, %s)",
                    (user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales)
                )
                conexion.commit()
                id_generado = cursor.lastrowid

        return respuesta_exitosa({"id": id_generado}, "Historial de recompensa registrado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al registrar historial de recompensa: {str(e)}")


@app.route("/api/historial_recompensas/actualizar", methods=['POST', 'PUT'])
@jwt_required()
def api_actualizar_historial_recompensa():
    """Actualiza un historial de recompensa existente (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_historial = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))
        user_id = datos.get("user_id")
        sesion_id = datos.get("sesion_id")
        tipo_sesion = datos.get("tipo_sesion")
        posicion = datos.get("posicion")
        puntosmoneda_ganados = datos.get("puntosmoneda_ganados")
        puntos_totales = datos.get("puntos_totales")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute(
                    "UPDATE historial_recompensas SET user_id=%s, sesion_id=%s, tipo_sesion=%s, posicion=%s, puntosmoneda_ganados=%s, puntos_totales=%s WHERE id=%s",
                    (user_id, sesion_id, tipo_sesion, posicion, puntosmoneda_ganados, puntos_totales, id_historial)
                )
                if filas == 0:
                    return respuesta_error("Historial de recompensa no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_historial}, "Historial de recompensa actualizado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al actualizar historial de recompensa: {str(e)}")


@app.route("/api/historial_recompensas/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_historial_recompensa_id(id=0):
    """Obtiene un historial de recompensa por ID"""
    try:
        target_id = id
        if request.method == 'POST':
            datos = obtener_datos_request()
            target_id = datos.get("id")
        try:
            target_id = validar_entero(target_id, "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM historial_recompensas WHERE id=%s", (target_id,))
                historial = cursor.fetchone()

        if historial:
            return respuesta_exitosa(historial, "Historial de recompensa obtenido correctamente")
        else:
            return respuesta_error("Historial de recompensa no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener historial de recompensa: {str(e)}")


@app.route("/api/historial_recompensas/obtener", methods=['GET'])
@jwt_required()
def api_obtener_historial_recompensas():
    """Obtiene todos los historiales de recompensas"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM historial_recompensas")
                historiales = cursor.fetchall()

        return respuesta_exitosa(historiales, f"Se encontraron {len(historiales)} historiales de recompensas")
    except Exception as e:
        return respuesta_error(f"Error al obtener historiales de recompensas: {str(e)}")


@app.route("/api/historial_recompensas/eliminar", methods=['POST', 'DELETE'])
@jwt_required()
def api_eliminar_historial_recompensa():
    """Elimina un historial de recompensa (Requiere JWT)"""
    try:
        datos = obtener_datos_request()
        try:
            id_historial = validar_entero(obtener_id_param(datos), "id")
        except ValueError as err:
            return respuesta_error(str(err))

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                filas = cursor.execute("DELETE FROM historial_recompensas WHERE id=%s", (id_historial,))
                if filas == 0:
                    return respuesta_error("Historial de recompensa no encontrado")
                conexion.commit()

        return respuesta_exitosa({"id": id_historial}, "Historial de recompensa eliminado correctamente")
    except Exception as e:
        return respuesta_error(f"Error al eliminar historial de recompensa: {str(e)}")


# =====================================================
# APIs VISTA: usuarios_con_rangos (VISTA - SOLO LECTURA)
# Esta es una VISTA calculada, no una tabla
# Estructura: id, username, email, puntosmoneda, rango_nombre, rango_icono,
#             rango_color, rango_orden, puntos_minimos, puntos_maximos,
#             puntos_para_siguiente_rango, siguiente_rango
# NOTA: Solo tiene operaciones GET (no se puede INSERT/UPDATE/DELETE en una vista)
# =====================================================


@app.route("/api/usuarios_con_rangos/obtener/<int:id>", methods=['GET', 'POST'])
@jwt_required()
def api_obtener_usuario_con_rango_id(id=0):
    """Obtiene información de un usuario con su rango asignado por ID"""
    try:
        if request.method == 'POST':
            id = request.json.get("id")

        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuarios_con_rangos WHERE id=%s", (id,))
                usuario_rango = cursor.fetchone()

        if usuario_rango:
            return respuesta_exitosa(usuario_rango, "Usuario con rango obtenido correctamente")
        else:
            return respuesta_error("Usuario con rango no encontrado")
    except Exception as e:
        return respuesta_error(f"Error al obtener usuario con rango: {str(e)}")


@app.route("/api/usuarios_con_rangos/obtener", methods=['GET'])
@jwt_required()
def api_obtener_usuarios_con_rangos():
    """Obtiene todos los usuarios con sus rangos asignados"""
    try:
        with bd.obtener_conexion_segura() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM usuarios_con_rangos")
                usuarios_rangos = cursor.fetchall()

        return respuesta_exitosa(usuarios_rangos, f"Se encontraron {len(usuarios_rangos)} usuarios con rangos")
    except Exception as e:
        return respuesta_error(f"Error al obtener usuarios con rangos: {str(e)}")



#--------------------------------------------




if __name__ == '__main__':
    app.run(debug=True)
